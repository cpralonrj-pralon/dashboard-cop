import io
import unittest

import openpyxl

from src.config import (
    EQUIPE_IDS,
    PRALON_ANALYSTS,
    RES_COL_GRUPO,
    RES_COL_IMPACTO,
    RES_COL_INDICADOR_NOME,
    RES_COL_LOGIN,
    RES_COL_LOGIN_FO,
    RES_COL_LOGIN_GPON,
    RES_COL_LOGIN_UNIFIED,
    RES_COL_SERVICO,
    RES_COL_SOLUCAO,
    RES_IND_ASSERT_FIBRA_HFC,
    RES_IND_ASSERT_GPON,
    RES_IND_ETIT_FIBRA_HFC,
    RES_IND_ETIT_GPON,
    RES_INDICADORES_FILTRO,
)
from src.processors import (
    ResidentialImportError,
    filter_residential_by_logins,
    load_residencial_indicadores,
    res_kpis_por_indicador,
    res_por_analista,
    res_por_solucao,
)


CANONICAL_HEADERS = [
    "INDICADOR_NOME_ICG",
    "ID_MOSTRA",
    "VOLUME",
    "INDICADOR",
    "INDICADOR_STATUS",
    "IN_REGIONAL",
    "IN_GRUPO",
    "SERVICO",
    "SOLUCAO",
    "IMPACTO",
    "DT_INICIO",
    "ANOMES",
]


def _row(indicator, login, adherent, service, impact, group, solution):
    return {
        "INDICADOR_NOME_ICG": indicator,
        "ID_MOSTRA": f"ID-{indicator}-{login}-{adherent}-{service}",
        "VOLUME": 1,
        "INDICADOR": 1 if adherent else 0,
        "INDICADOR_STATUS": "ADERENTE" if adherent else "NÃO ADERENTE",
        "IN_REGIONAL": "Leste",
        "IN_GRUPO": group,
        "SERVICO": service,
        "SOLUCAO": solution,
        "IMPACTO": impact,
        "DT_INICIO": "2026-08-10 23:30:00",
        "ANOMES": 202608,
        "_LOGIN": login,
    }


def _fixture_rows():
    team = ["N5972428", "N5577565"]
    rows = []
    for index, indicator in enumerate(RES_INDICADORES_FILTRO):
        is_gpon = indicator in {RES_IND_ETIT_GPON, RES_IND_ASSERT_GPON}
        rows.append(
            _row(
                indicator,
                team[0],
                True,
                " greenfield " if is_gpon else "HFC",
                "Massivo",
                "Rio e ES",
                "SOLUÇÃO A",
            )
        )
        rows.append(
            _row(
                indicator.lower(),
                team[1],
                False,
                "Brownfield" if is_gpon else "HFC",
                " NAO MASSIVO ",
                "Centro-Oeste",
                "SOLUÇÃO B",
            )
        )
    rows.append(
        _row(
            RES_IND_ETIT_GPON,
            "FORA_DA_EQUIPE",
            True,
            "GREENFIELD",
            "Massivo",
            "Norte",
            "SOLUÇÃO EXTERNA",
        )
    )
    return rows


def _workbook_bytes(login_schema="unified", varied_headers=False, preamble_rows=0, rows=None):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = " analítico " if varied_headers else "Analitico"

    for _ in range(preamble_rows):
        worksheet.append(["Relatório Residencial 202608"])

    headers = list(CANONICAL_HEADERS)
    if login_schema == "unified":
        headers.append(RES_COL_LOGIN_UNIFIED)
    elif login_schema == "legacy":
        headers.extend([RES_COL_LOGIN_FO, RES_COL_LOGIN_GPON])
    else:
        raise ValueError(login_schema)

    if varied_headers:
        display_headers = [
            " indicador nome icg ",
            "id mostra",
            "volume",
            "indicador",
            "indicador status",
            "in regional",
            "in grupo",
            "serviço",
            "solução",
            "impacto",
            "dt início",
            "anomes",
        ]
        if login_schema == "unified":
            display_headers.append(" login primeiro acionamento ")
        else:
            display_headers.extend(
                ["login primeiro acionamento fo", "login primeiro acionamento gpon"]
            )
        worksheet.append(display_headers)
    else:
        worksheet.append(headers)

    for data in rows or _fixture_rows():
        values = [data.get(header) for header in CANONICAL_HEADERS]
        if login_schema == "unified":
            values.append(data.get("_LOGIN"))
        else:
            is_hfc = str(data["INDICADOR_NOME_ICG"]).strip().upper() in {
                RES_IND_ETIT_FIBRA_HFC,
                RES_IND_ASSERT_FIBRA_HFC,
            }
            values.extend(
                [data.get("_LOGIN") if is_hfc else None, data.get("_LOGIN") if not is_hfc else None]
            )
        worksheet.append(values)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class ResidentialImportTest(unittest.TestCase):
    def test_current_unified_login_schema_finds_nelson_team(self):
        imported = load_residencial_indicadores(_workbook_bytes("unified"))
        team = filter_residential_by_logins(imported, EQUIPE_IDS)

        self.assertEqual(9, len(imported))
        self.assertEqual(8, len(team))
        self.assertEqual({"N5972428", "N5577565"}, set(team[RES_COL_LOGIN]))
        self.assertEqual(0, team[RES_COL_LOGIN].eq("").sum())

    def test_pralon_scope_uses_its_own_permission_map(self):
        imported = load_residencial_indicadores(_workbook_bytes("unified"))
        scoped = filter_residential_by_logins(imported, PRALON_ANALYSTS)

        self.assertEqual(8, len(scoped))
        self.assertNotIn("FORA_DA_EQUIPE", set(scoped[RES_COL_LOGIN]))

    def test_legacy_fo_and_gpon_login_columns_remain_supported(self):
        imported = load_residencial_indicadores(_workbook_bytes("legacy"))
        team = filter_residential_by_logins(imported, EQUIPE_IDS)

        self.assertEqual(8, len(team))
        self.assertEqual({"N5972428", "N5577565"}, set(team[RES_COL_LOGIN]))
        self.assertEqual(
            sorted([RES_COL_LOGIN_FO, RES_COL_LOGIN_GPON]),
            imported.attrs["residential_import_audit"]["login_source_columns"],
        )

    def test_header_variations_and_preamble_are_normalized(self):
        imported = load_residencial_indicadores(
            _workbook_bytes("unified", varied_headers=True, preamble_rows=2)
        )

        self.assertEqual(9, len(imported))
        self.assertEqual(3, imported.attrs["residential_import_audit"]["header_row"])
        self.assertIn(RES_COL_SOLUCAO, imported.columns)
        self.assertIn(RES_COL_LOGIN_UNIFIED, imported.columns)

    def test_four_indicators_and_all_required_classifications(self):
        imported = load_residencial_indicadores(_workbook_bytes("unified"))
        team = filter_residential_by_logins(imported, EQUIPE_IDS)
        kpis = res_kpis_por_indicador(team).set_index("Indicador")

        self.assertEqual(set(RES_INDICADORES_FILTRO), set(kpis.index))
        for indicator in RES_INDICADORES_FILTRO:
            self.assertEqual(2, int(kpis.loc[indicator, "Volume"]))
            self.assertEqual(1, int(kpis.loc[indicator, "Aderentes"]))
            self.assertEqual(1, int(kpis.loc[indicator, "Nao_Aderentes"]))
            self.assertEqual(50.0, float(kpis.loc[indicator, "Aderencia_Pct"]))
            self.assertEqual(50.0, float(kpis.loc[indicator, "Nao_Aderencia_Pct"]))

        gpon = team[team[RES_COL_INDICADOR_NOME] == RES_IND_ETIT_GPON]
        self.assertEqual({"GREENFIELD", "BROWNFIELD"}, set(gpon[RES_COL_SERVICO]))
        self.assertEqual({"Massivo", "Não Massivo"}, set(gpon[RES_COL_IMPACTO]))
        self.assertEqual({"Rio e ES", "Centro-Oeste"}, set(gpon[RES_COL_GRUPO]))
        self.assertEqual({"SOLUÇÃO A", "SOLUÇÃO B"}, set(gpon[RES_COL_SOLUCAO]))

    def test_rankings_and_solutions_use_real_adherence(self):
        imported = load_residencial_indicadores(_workbook_bytes("unified"))
        team = filter_residential_by_logins(imported, EQUIPE_IDS)
        ranking = res_por_analista(team, RES_IND_ETIT_GPON)
        solutions = res_por_solucao(team, RES_IND_ETIT_GPON)

        self.assertEqual("N5972428", ranking.iloc[0]["Login"])
        self.assertEqual(100.0, float(ranking.iloc[0]["Aderencia_Pct"]))
        self.assertEqual("N5577565", ranking.iloc[-1]["Login"])
        self.assertEqual(100.0, float(ranking.iloc[-1]["Nao_Aderencia_Pct"]))
        self.assertEqual(2, int(solutions["Volume"].sum()))

    def test_valid_file_without_matching_team_is_a_filter_result(self):
        only_external = [
            _row(
                RES_IND_ETIT_GPON,
                "FORA_DA_EQUIPE",
                True,
                "GREENFIELD",
                "Massivo",
                "Norte",
                "SOLUÇÃO EXTERNA",
            )
        ]
        imported = load_residencial_indicadores(
            _workbook_bytes("unified", rows=only_external)
        )

        self.assertEqual(1, len(imported))
        self.assertTrue(filter_residential_by_logins(imported, EQUIPE_IDS).empty)

    def test_valid_legacy_minimal_schema_keeps_optional_fields_optional(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Analitico"
        worksheet.append(
            [
                "INDICADOR_NOME_ICG",
                "VOLUME",
                "INDICADOR",
                "IN_REGIONAL",
                RES_COL_LOGIN_FO,
            ]
        )
        worksheet.append([RES_IND_ETIT_FIBRA_HFC, "1", "1", "LESTE", "n5972428"])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        imported = load_residencial_indicadores(payload)
        audit = imported.attrs["residential_import_audit"]

        self.assertEqual(1, len(imported))
        self.assertEqual("N5972428", imported.iloc[0][RES_COL_LOGIN])
        self.assertIn(RES_COL_GRUPO, audit["missing_optional_columns"])
        self.assertIn(RES_COL_SOLUCAO, audit["missing_optional_columns"])

    def test_invalid_schema_is_not_reported_as_empty_filter(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Analitico"
        worksheet.append(["INDICADOR_NOME_ICG", RES_COL_LOGIN_UNIFIED])
        worksheet.append([RES_IND_ETIT_GPON, "N5972428"])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        with self.assertRaises(ResidentialImportError) as context:
            load_residencial_indicadores(payload)

        self.assertIn(
            context.exception.code,
            {"schema_incompatible", "required_column_missing"},
        )


if __name__ == "__main__":
    unittest.main()
