import pandas as pd
import numpy as np
import openpyxl
import io
from src.config import (
    EQUIPE_IDS, ALL_TRACKED_IDS, BASE_EQUIPE, HEADER_ROW, SHEET_NAME_CANDIDATES,
    REGIONAL_FILTRO, LOGIN_ALIASES, COORD_ANALYSTS_NAMES,
    SUB_ADMIN_EMP_IDS, COORD_ANALYSTS_MAP,
    COL_LOGIN, COL_NOME, COL_BASE, COL_DATA, COL_MES, COL_ANOMES,
    COL_VOL_TOTAL, COL_VOL_MEDIA, COL_DPA_USO, COL_DPA_JORNADA,
    COL_DPA_RESULTADO, VOL_COLS, COL_CARGO, COL_PERIODO, COL_COORD,
    # ETIT
    ETIT_INDICADOR_FILTRO, ETIT_COL_INDICADOR, ETIT_COL_LOGIN,
    ETIT_COL_DEMANDA, ETIT_COL_VOLUME, ETIT_COL_STATUS,
    ETIT_COL_TIPO, ETIT_COL_AREA, ETIT_COL_CAUSA,
    ETIT_COL_REGIONAL, ETIT_COL_GRUPO, ETIT_COL_CIDADE, ETIT_COL_UF,
    ETIT_COL_TOA, ETIT_COL_DT_INICIO, ETIT_COL_DT_FIM,
    ETIT_COL_DT_ACIONAMENTO, ETIT_COL_TURNO,
    ETIT_COL_TMA, ETIT_COL_TMR, ETIT_COL_ANOMES,
    ETIT_SHEET_CANDIDATES, ETIT_COL_INDICADOR_VAL,
    # Residencial Indicadores
    RES_INDICADORES_FILTRO, RES_IND_INVERTIDOS,
    RES_IND_ETIT_FIBRA_HFC, RES_IND_ETIT_GPON, RES_IND_LOG_REPROG_GPON,
    RES_IND_ASSERT_FIBRA_HFC, RES_IND_ASSERT_GPON,
    RES_SHEET_CANDIDATES,
    RES_COL_INDICADOR_NOME, RES_COL_ID_MOSTRA, RES_COL_VOLUME,
    RES_COL_LOGIN_FO, RES_COL_LOGIN_GPON, RES_COL_LOGIN,
    RES_COL_INDICADOR_VAL, RES_COL_STATUS, RES_COL_REGIONAL,
    RES_COL_GRUPO, RES_COL_CIDADE, RES_COL_UF, RES_COL_TECNOLOGIA,
    RES_COL_SERVICO, RES_COL_NATUREZA, RES_COL_SINTOMA,
    RES_COL_FERRAMENTA, RES_COL_FECHAMENTO, RES_COL_SOLUCAO,
    RES_COL_IMPACTO, RES_COL_ENVIADO_TOA, RES_COL_DT_INICIO,
    RES_COL_DT_FIM, RES_COL_DT_FIM_SISTEMA, RES_COL_TMA, RES_COL_TMR, RES_COL_ANOMES, RES_COL_TURNO,
    # DPA Ocupação
    DPA_MESES_PT, DPA_SHEET_ANALISTAS, DPA_SHEET_CONSOLIDADO,
    # Indicadores TOA
    TOA_IND_SHEET, TOA_INDICADORES_FILTRO, TOA_IND_INVERTIDOS,
    TOA_COL_INDICADOR_NOME, TOA_COL_LOGIN, TOA_COL_INDICADOR,
    TOA_COL_STATUS, TOA_COL_REGIONAL,
    TOA_COL_TIPO_ATIVIDADE, TOA_COL_REDE, TOA_COL_MERCADO,
    TOA_COL_NATUREZA, TOA_COL_SOLUCAO,
    TOA_COL_TMR, TOA_COL_AGING, TOA_COL_DATA,
    TOA_COL_DT_CANCELAMENTO, TOA_COL_DT_INICIO_FORM, TOA_COL_DT_FIM_FORM,
    TOA_COL_ANOMES, TOA_COL_ID_ATIVIDADE, TOA_COL_RESPONSAVEL, TOA_AGING_ORDER,
    TOA_IND_CANCELADAS, TOA_IND_VALIDACAO,
)

# Lookup: matrícula → "EMPRESARIAL" para analistas das equipes sub-admin empresarial.
# Garante que o filtro de setor funcione corretamente para N0150817/N5768308/TPAROLI.
_COORD_EMP_ANALYST_SECTOR: dict[str, str] = {
    m.upper(): "EMPRESARIAL"
    for cid in SUB_ADMIN_EMP_IDS
    for m in COORD_ANALYSTS_MAP.get(cid, set())
}


def list_sheets(uploaded_file):
    data = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    return wb.sheetnames


def load_produtividade(uploaded_file) -> pd.DataFrame:
    """Lê a planilha de produtividade e retorna DataFrame filtrado pela equipe."""
    sheets = list_sheets(uploaded_file)

    sheet_to_read = None
    for candidate in SHEET_NAME_CANDIDATES:
        if candidate in sheets:
            sheet_to_read = candidate
            break
    if sheet_to_read is None:
        sheet_to_read = sheets[0]

    df = pd.read_excel(uploaded_file, sheet_name=sheet_to_read, header=HEADER_ROW)

    # Remove coluna unnamed (colunas com cabecalho vazio/numerico precisam ser
    # forcadas para string antes do startswith para nao quebrar com NaN).
    _col_str = df.columns.astype(str)
    df = df.loc[:, ~_col_str.str.startswith("Unnamed")]

    if COL_LOGIN not in df.columns:
        raise ValueError(
            f"Coluna '{COL_LOGIN}' nao encontrada na aba '{sheet_to_read}' "
            f"(cabecalho esperado na linha {HEADER_ROW + 1}). "
            f"Verifique se o arquivo enviado e a planilha de Produtividade correta. "
            f"Colunas lidas: {list(df.columns)[:25]}"
        )

    # Normaliza matrículas alternativas antes de filtrar pela equipe
    df[COL_LOGIN] = (
        df[COL_LOGIN].astype(str).str.strip()
        .map(lambda x, a=LOGIN_ALIASES: a.get(x, x))
    )
    df_equipe = df[df[COL_LOGIN].isin(ALL_TRACKED_IDS)].copy()

    # Garante tipos numéricos
    num_cols = [COL_VOL_TOTAL, COL_VOL_MEDIA, COL_DPA_USO, COL_DPA_JORNADA, COL_DPA_RESULTADO]
    num_cols += list(VOL_COLS.keys())
    for c in num_cols:
        if c in df_equipe.columns:
            df_equipe[c] = pd.to_numeric(df_equipe[c], errors="coerce")

    # Data como datetime
    if COL_DATA in df_equipe.columns:
        df_equipe[COL_DATA] = pd.to_datetime(df_equipe[COL_DATA], errors="coerce")

    # Merge com info da equipe (setor fixo)
    df_equipe = df_equipe.merge(
        BASE_EQUIPE[["Matricula", "Setor"]],
        left_on=COL_LOGIN, right_on="Matricula", how="left"
    )
    df_equipe["Setor"] = df_equipe["Setor"].fillna(
        df_equipe[COL_LOGIN].str.upper().map(_COORD_EMP_ANALYST_SECTOR)
    ).fillna("")

    return df_equipe


# =====================================================
# ETIT POR EVENTO — Loader e processadores
# =====================================================
def load_etit(uploaded_file) -> pd.DataFrame:
    """Lê a planilha Analítico Empresarial e retorna apenas ETIT POR EVENTO da equipe."""
    sheets = list_sheets(uploaded_file)
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    sheet_to_read = None
    for candidate in ETIT_SHEET_CANDIDATES:
        if candidate in sheets:
            sheet_to_read = candidate
            break
    if sheet_to_read is None:
        sheet_to_read = sheets[0]

    df = pd.read_excel(uploaded_file, sheet_name=sheet_to_read)

    if ETIT_COL_INDICADOR not in df.columns:
        raise ValueError(
            f"Coluna '{ETIT_COL_INDICADOR}' não encontrada na aba '{sheet_to_read}'. "
            f"Abas disponíveis: {sheets}. Colunas lidas: {list(df.columns)[:20]}..."
        )

    df = df[df[ETIT_COL_INDICADOR] == ETIT_INDICADOR_FILTRO].copy()
    if df.empty:
        raise ValueError(
            f"Nenhuma linha com {ETIT_COL_INDICADOR} == '{ETIT_INDICADOR_FILTRO}' "
            f"na aba '{sheet_to_read}'. Verifique se a planilha contém este indicador."
        )

    # Normaliza matrículas alternativas antes de filtrar pela equipe
    df[ETIT_COL_LOGIN] = (
        df[ETIT_COL_LOGIN].astype(str).str.strip()
        .map(lambda x, a=LOGIN_ALIASES: a.get(x, x))
    )
    df_equipe = df[df[ETIT_COL_LOGIN].isin(ALL_TRACKED_IDS)].copy()

    # Filtra regional Leste
    if ETIT_COL_REGIONAL in df_equipe.columns:
        df_equipe = df_equipe[df_equipe[ETIT_COL_REGIONAL] == REGIONAL_FILTRO].copy()

    if df_equipe.empty:
        raise ValueError(
            f"Após filtrar por equipe (ALL_TRACKED_IDS) e regional '{REGIONAL_FILTRO}', "
            f"nenhum registro restou. Verifique se a planilha contém analistas da equipe "
            f"e a regional Leste."
        )

    # Merge com info da equipe (nome e setor)
    df_equipe = df_equipe.merge(
        BASE_EQUIPE[["Matricula", "Nome", "Setor"]],
        left_on=ETIT_COL_LOGIN, right_on="Matricula", how="left"
    )
    df_equipe["Nome"] = df_equipe["Nome"].fillna(
        df_equipe[ETIT_COL_LOGIN].str.upper().map(COORD_ANALYSTS_NAMES)
    ).fillna(df_equipe[ETIT_COL_LOGIN])
    df_equipe["Setor"] = df_equipe["Setor"].fillna(
        df_equipe[ETIT_COL_LOGIN].str.upper().map(_COORD_EMP_ANALYST_SECTOR)
    ).fillna("")

    # Garante tipos
    for c in [ETIT_COL_TMA, ETIT_COL_TMR, ETIT_COL_VOLUME, ETIT_COL_INDICADOR_VAL]:
        if c in df_equipe.columns:
            df_equipe[c] = pd.to_numeric(df_equipe[c], errors="coerce")

    for c in [ETIT_COL_DT_INICIO, ETIT_COL_DT_FIM, ETIT_COL_DT_ACIONAMENTO]:
        if c in df_equipe.columns:
            df_equipe[c] = pd.to_datetime(df_equipe[c], errors="coerce")

    if ETIT_COL_ANOMES in df_equipe.columns:
        df_equipe[ETIT_COL_ANOMES] = (
            df_equipe[ETIT_COL_ANOMES]
            .astype(str).str.strip()
            .str.replace(r"\.0+$", "", regex=True)
        )

    return df_equipe


def etit_resumo_analista(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    group = df.groupby([ETIT_COL_LOGIN, "Nome", "Setor"]).agg(
        Total_Eventos=(ETIT_COL_VOLUME, "sum"),
        Eventos_Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
        TMA_Medio=(ETIT_COL_TMA, "mean"),
        TMR_Medio=(ETIT_COL_TMR, "mean"),
        RAL_Count=(ETIT_COL_DEMANDA, lambda x: (x == "RAL").sum()),
        REC_Count=(ETIT_COL_DEMANDA, lambda x: (x == "REC").sum()),
    ).reset_index()
    group["Aderencia_Pct"] = (group["Eventos_Aderentes"] / group["Total_Eventos"] * 100).round(1)
    group["TMA_Medio"] = group["TMA_Medio"].round(4)
    group["TMR_Medio"] = group["TMR_Medio"].round(4)
    return group.sort_values("Total_Eventos", ascending=False).reset_index(drop=True)


def etit_por_demanda(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(ETIT_COL_DEMANDA).agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
        TMA_Medio=(ETIT_COL_TMA, "mean"),
        TMR_Medio=(ETIT_COL_TMR, "mean"),
    ).reset_index().rename(columns={ETIT_COL_DEMANDA: "Demanda"})


def etit_por_tipo(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(ETIT_COL_TIPO).agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
    ).reset_index().rename(columns={ETIT_COL_TIPO: "Tipo"}).sort_values("Eventos", ascending=False)


def etit_por_causa(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(ETIT_COL_CAUSA).agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
    ).reset_index().rename(columns={ETIT_COL_CAUSA: "Causa"}).sort_values("Eventos", ascending=False)


def etit_por_regional(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(ETIT_COL_REGIONAL).agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
    ).reset_index().rename(columns={ETIT_COL_REGIONAL: "Regional"}).sort_values("Eventos", ascending=False)


def etit_por_turno(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(ETIT_COL_TURNO).agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
    ).reset_index().rename(columns={ETIT_COL_TURNO: "Turno"}).sort_values("Eventos", ascending=False)


def etit_aderencia_ral_rec_por_analista(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna aderentes e não aderentes por analista, separados por RAL e REC."""
    if df.empty or ETIT_COL_DEMANDA not in df.columns:
        return pd.DataFrame()

    grp = df.groupby([ETIT_COL_LOGIN, "Nome", ETIT_COL_DEMANDA]).agg(
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
        Total=(ETIT_COL_VOLUME, "sum"),
    ).reset_index()
    grp["Nao_Aderentes"] = (grp["Total"] - grp["Aderentes"]).clip(lower=0).astype(int)
    grp["Aderentes"] = grp["Aderentes"].astype(int)

    pivot = grp.pivot_table(
        index=[ETIT_COL_LOGIN, "Nome"],
        columns=ETIT_COL_DEMANDA,
        values=["Aderentes", "Nao_Aderentes"],
        fill_value=0,
    ).reset_index()

    # Flatten multi-level columns: (metric, DEMANDA) → "DEMANDA_metric"
    flat_cols = []
    for col in pivot.columns:
        if col[1] == "":
            flat_cols.append(col[0])
        else:
            flat_cols.append(f"{col[1]}_{col[0]}")
    pivot.columns = flat_cols

    ader_cols = [c for c in pivot.columns if c.endswith("_Aderentes")]
    pivot["_sort"] = pivot[ader_cols].sum(axis=1)
    pivot = pivot.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    return pivot.reset_index(drop=True)


def etit_evolucao_diaria(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or ETIT_COL_DT_ACIONAMENTO not in df.columns:
        return pd.DataFrame()
    df_c = df.copy()
    df_c["Data"] = df_c[ETIT_COL_DT_ACIONAMENTO].dt.date
    daily = df_c.groupby("Data").agg(
        Eventos=(ETIT_COL_VOLUME, "sum"),
        Aderentes=(ETIT_COL_INDICADOR_VAL, "sum"),
        Analistas=(ETIT_COL_LOGIN, "nunique"),
    ).reset_index()
    daily["Data"] = pd.to_datetime(daily["Data"])
    daily["Aderencia_Pct"] = (daily["Aderentes"] / daily["Eventos"] * 100).round(1)
    return daily


# =====================================================
# INDICADORES RESIDENCIAL — Loader e processadores
# =====================================================

def load_residencial_indicadores(uploaded_file) -> pd.DataFrame:
    sheets = list_sheets(uploaded_file)
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    sheet_to_read = None
    for candidate in RES_SHEET_CANDIDATES:
        if candidate in sheets:
            sheet_to_read = candidate
            break
    if sheet_to_read is None:
        sheet_to_read = sheets[0]

    df = pd.read_excel(uploaded_file, sheet_name=sheet_to_read)

    if RES_COL_INDICADOR_NOME not in df.columns:
        raise ValueError(
            f"Coluna '{RES_COL_INDICADOR_NOME}' não encontrada na aba '{sheet_to_read}'. "
            f"Abas disponíveis: {sheets}. Colunas lidas: {list(df.columns)[:20]}..."
        )

    df = df[df[RES_COL_INDICADOR_NOME].isin(RES_INDICADORES_FILTRO)].copy()
    if df.empty:
        raise ValueError(
            f"Nenhuma linha com {RES_COL_INDICADOR_NOME} em {RES_INDICADORES_FILTRO} "
            f"na aba '{sheet_to_read}'. Verifique se a planilha contém estes indicadores."
        )

    # Filtra regional Leste
    if RES_COL_REGIONAL in df.columns:
        df = df[df[RES_COL_REGIONAL] == REGIONAL_FILTRO].copy()

    if df.empty:
        raise ValueError(
            f"Após filtrar por regional '{REGIONAL_FILTRO}' (coluna {RES_COL_REGIONAL}), "
            f"nenhum registro restou."
        )

    # Normaliza matrículas alternativas para a matrícula canônica
    if LOGIN_ALIASES:
        for _col in [RES_COL_ID_MOSTRA, "LOGIN_ACIONAMENTO", "LOGIN"]:
            if _col in df.columns:
                df[_col] = (
                    df[_col].astype(str).str.strip()
                    .map(lambda x, a=LOGIN_ALIASES: a.get(x, x))
                )

    for c in [RES_COL_VOLUME, RES_COL_INDICADOR_VAL, RES_COL_TMA, RES_COL_TMR]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    for c in [RES_COL_DT_INICIO, RES_COL_DT_FIM, RES_COL_DT_FIM_SISTEMA]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    # Calcula TURNO a partir do horário de referência de cada indicador:
    # ETIT Fibra HFC / ETIT GPON → DT_INICIO
    # LOG REPROG GPON             → DT_FIM_SISTEMA_PRIMEIRO_FECHAMENTO
    def _hora_para_turno(h):
        if pd.isna(h):
            return "Madrugada"
        h = int(h)
        if 6 <= h <= 13:
            return "Manhã"
        elif 14 <= h <= 21:
            return "Tarde"
        return "Madrugada"

    etit_mask   = df[RES_COL_INDICADOR_NOME].isin({RES_IND_ETIT_FIBRA_HFC, RES_IND_ETIT_GPON, RES_IND_ASSERT_FIBRA_HFC, RES_IND_ASSERT_GPON})
    reprog_mask = df[RES_COL_INDICADOR_NOME] == RES_IND_LOG_REPROG_GPON

    turno_series = pd.Series("Madrugada", index=df.index, dtype=str)
    if RES_COL_DT_INICIO in df.columns:
        turno_series[etit_mask] = (
            df.loc[etit_mask, RES_COL_DT_INICIO].dt.hour.map(_hora_para_turno)
        )
    if RES_COL_DT_FIM_SISTEMA in df.columns:
        turno_series[reprog_mask] = (
            df.loc[reprog_mask, RES_COL_DT_FIM_SISTEMA].dt.hour.map(_hora_para_turno)
        )
    df[RES_COL_TURNO] = turno_series

    if RES_COL_ANOMES in df.columns:
        df[RES_COL_ANOMES] = (
            df[RES_COL_ANOMES]
            .astype(str).str.strip()
            .str.replace(r"\.0+$", "", regex=True)
        )

    # ADERENTE:
    # - LOG REPROGRAMAÇÃO GPON: INDICADOR=0 → aderente (sem reprogramação); 1+ → não aderente
    # - Demais: INDICADOR=1 → aderente
    df["ADERENTE"] = df.apply(
        lambda row: (
            (row[RES_COL_INDICADOR_VAL] == 0)
            if row[RES_COL_INDICADOR_NOME] in RES_IND_INVERTIDOS
            else (row[RES_COL_INDICADOR_VAL] == 1)
        ),
        axis=1,
    ).astype(int)

    # QTDE_REPROG: para LOG REPROGRAMAÇÃO GPON guarda o número de reprogramações (valor do INDICADOR)
    # Para os demais indicadores o campo fica None
    df["QTDE_REPROG"] = df.apply(
        lambda row: row[RES_COL_INDICADOR_VAL]
        if row[RES_COL_INDICADOR_NOME] == RES_IND_LOG_REPROG_GPON
        else None,
        axis=1,
    )

    if RES_COL_DT_INICIO in df.columns:
        df["DATA_DIA"] = df[RES_COL_DT_INICIO].dt.normalize()

    # Coluna unificada de login: cada indicador usa sua própria coluna de origem
    fo_mask   = df[RES_COL_INDICADOR_NOME].isin({RES_IND_ETIT_FIBRA_HFC, RES_IND_ASSERT_FIBRA_HFC})
    gpon_mask = ~fo_mask

    df[RES_COL_LOGIN] = ""
    if RES_COL_LOGIN_FO in df.columns:
        df.loc[fo_mask,   RES_COL_LOGIN] = (
            df.loc[fo_mask,   RES_COL_LOGIN_FO]
            .astype(str).str.strip().str.upper()
        )
    if RES_COL_LOGIN_GPON in df.columns:
        df.loc[gpon_mask, RES_COL_LOGIN] = (
            df.loc[gpon_mask, RES_COL_LOGIN_GPON]
            .astype(str).str.strip().str.upper()
        )

    # Normaliza aliases e remove "nan" resultante de células vazias
    if LOGIN_ALIASES:
        df[RES_COL_LOGIN] = df[RES_COL_LOGIN].map(lambda x: LOGIN_ALIASES.get(x, x))
    df[RES_COL_LOGIN] = df[RES_COL_LOGIN].replace("NAN", "")

    # Merge com BASE_EQUIPE para nome e setor
    df = df.merge(
        BASE_EQUIPE[["Matricula", "Nome", "Setor"]],
        left_on=RES_COL_LOGIN, right_on="Matricula", how="left",
    ).drop(columns="Matricula", errors="ignore")
    df["Nome"] = df["Nome"].fillna(
        df[RES_COL_LOGIN].map(COORD_ANALYSTS_NAMES)
    ).fillna(df[RES_COL_LOGIN])
    df["Setor"] = df["Setor"].fillna("")

    return df


def res_kpis_por_indicador(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    has_tma    = RES_COL_TMA in df.columns
    has_tmr    = RES_COL_TMR in df.columns
    has_reprog = "QTDE_REPROG" in df.columns
    agg = {RES_COL_VOLUME: "sum", "ADERENTE": "sum"}
    if has_tma:
        agg[RES_COL_TMA] = "mean"
    if has_tmr:
        agg[RES_COL_TMR] = "mean"
    if has_reprog:
        agg["QTDE_REPROG"] = "sum"
    g = df.groupby(RES_COL_INDICADOR_NOME).agg(agg).reset_index()
    col_names = (
        ["Indicador", "Volume", "Aderentes"]
        + (["TMA_Medio"] if has_tma else [])
        + (["TMR_Medio"] if has_tmr else [])
        + (["QTDE_REPROG"] if has_reprog else [])
    )
    g.columns = col_names
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    order = {ind: i for i, ind in enumerate(RES_INDICADORES_FILTRO)}
    g["_ord"] = g["Indicador"].map(order)
    return g.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)


def res_por_analista(df: pd.DataFrame, indicador: str = None) -> pd.DataFrame:
    """Ranking de aderência por analista para um indicador específico."""
    if df.empty or RES_COL_LOGIN not in df.columns:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador].copy() if indicador else df.copy()
    sub = sub[sub[RES_COL_LOGIN] != ""]
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby([RES_COL_LOGIN, "Nome", "Setor"]).agg(
        Volume=(RES_COL_VOLUME, "sum"),
        Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={RES_COL_LOGIN: "Login"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Aderencia_Pct", ascending=False).reset_index(drop=True)


def res_por_regional(df: pd.DataFrame, indicador=None) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador] if indicador else df
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby(RES_COL_REGIONAL).agg(
        Volume=(RES_COL_VOLUME, "sum"), Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={RES_COL_REGIONAL: "Regional"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).reset_index(drop=True)


def res_por_natureza(df: pd.DataFrame, indicador=None) -> pd.DataFrame:
    if df.empty or RES_COL_NATUREZA not in df.columns:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador] if indicador else df
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby(RES_COL_NATUREZA).agg(
        Volume=(RES_COL_VOLUME, "sum"), Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={RES_COL_NATUREZA: "Natureza"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).reset_index(drop=True)


def res_por_solucao(df: pd.DataFrame, indicador=None, top_n=15) -> pd.DataFrame:
    if df.empty or RES_COL_SOLUCAO not in df.columns:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador] if indicador else df
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby(RES_COL_SOLUCAO).agg(
        Volume=(RES_COL_VOLUME, "sum"), Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={RES_COL_SOLUCAO: "Solução"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).head(top_n).reset_index(drop=True)


def res_por_impacto(df: pd.DataFrame, indicador=None) -> pd.DataFrame:
    if df.empty or RES_COL_IMPACTO not in df.columns:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador] if indicador else df
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby(RES_COL_IMPACTO).agg(
        Volume=(RES_COL_VOLUME, "sum"), Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={RES_COL_IMPACTO: "Impacto"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).reset_index(drop=True)


def res_evolucao_diaria(df: pd.DataFrame, indicador=None) -> pd.DataFrame:
    if df.empty or "DATA_DIA" not in df.columns:
        return pd.DataFrame()
    sub = df[df[RES_COL_INDICADOR_NOME] == indicador] if indicador else df
    if sub.empty:
        return pd.DataFrame()
    sub = sub.dropna(subset=["DATA_DIA"])
    g = sub.groupby("DATA_DIA").agg(
        Volume=(RES_COL_VOLUME, "sum"), Aderentes=("ADERENTE", "sum"),
    ).reset_index().rename(columns={"DATA_DIA": "Data"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Data").reset_index(drop=True)


# =====================================================
# OCUPAÇÃO DPA — Loader (planilha Ocupação_DPA_2026)
# =====================================================

def _dpa_detect_mes_recente(df_raw: pd.DataFrame) -> dict:
    """
    Varre todas as células da aba Consolidado buscando nome de mês PT.
    Quando encontra, pega a última coluna numérica decimal (0 < v <= 2) da
    mesma linha como % DPA.

    Problema: a planilha preenche todos os 12 meses (incluindo futuros) com
    valores projetados na mesma coluna, fazendo Novembro/Dezembro sempre
    vencer na busca pelo "mês mais recente".

    Solução: filtrar candidatos para meses ≤ mês atual do calendário — meses
    futuros só têm projeções, nunca dados reais. O maior mês passado com dado
    válido é o mês mais recente com dados reais.
    Fallback: se nenhum mês passado for encontrado, usa todos os candidatos.
    """
    import datetime
    current_month = datetime.date.today().month

    candidatos = []   # (mes_nome, mes_num, pct_f, pct_col_used)

    for _, row in df_raw.iterrows():
        # Procura mês em qualquer coluna da linha
        mes_val = None
        mes_col = None
        for col_idx, val in row.items():
            if str(val).strip() in DPA_MESES_PT:
                mes_val = str(val).strip()
                mes_col = col_idx
                break
        if mes_val is None:
            continue
        # Pega a última coluna numérica após o mês (% em decimal 0-2)
        pct_f = None
        pct_col_used = -1
        for col_idx, val in row.items():
            if col_idx <= mes_col:
                continue
            try:
                v = float(val)
                if 0 < v <= 2:
                    pct_f = v
                    pct_col_used = col_idx
            except (TypeError, ValueError):
                pass
        if pct_f is not None:
            mes_num = DPA_MESES_PT.index(mes_val) + 1
            candidatos.append((mes_val, mes_num, pct_f, pct_col_used))

    if not candidatos:
        return {"mes_nome": None, "mes_num": None, "dpa_geral_pct": None}

    # Filtrar apenas meses já ocorridos (≤ mês atual) — excluem projeções futuras
    pool = [c for c in candidatos if c[1] <= current_month] or candidatos

    # Dentre o pool: coluna mais à direita (dados do ano corrente), mês mais alto
    melhor_col = max(c[3] for c in pool)
    melhor = max((c for c in pool if c[3] == melhor_col), key=lambda c: c[1])

    return {
        "mes_nome": melhor[0],
        "mes_num": melhor[1],
        "dpa_geral_pct": round(melhor[2] * 100, 2),
    }


def _parse_pct_value(val) -> "float | None":
    """Parseia valor de % DPA: aceita float decimal (0.9275) ou string BR ('92,75%').
    Retorna float no intervalo 0-1, ou None se inválido/NaN."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        v = float(val)
        return None if pd.isna(v) else v
    except (TypeError, ValueError):
        pass
    # Tenta string com locale BR: "92,75%" → 0.9275
    s = str(val).strip().replace("%", "").replace(",", ".").strip()
    try:
        v = float(s)
        return v / 100 if v > 2 else v
    except (TypeError, ValueError):
        return None


def _dpa_extract_analistas(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extrai tabela de DPA por analista da aba Analistas (header=None).
    Detecta dinamicamente a linha de header, a coluna de Login e a coluna
    de % Produtivo do mês mais recente (última coluna com "%" no header).
    """
    header_row_idx = None
    login_col = None

    # Procura "Rótulos de Linha" em qualquer coluna
    for i, row in df_raw.iterrows():
        for col_idx, val in row.items():
            if str(val).strip() == "Rótulos de Linha":
                header_row_idx = i
                login_col = col_idx
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return pd.DataFrame()

    # Na linha de header, acha a PRIMEIRA coluna com "%" depois de login_col.
    # A tabela de analistas fica à esquerda; uma segunda tabela de dias do mês
    # pode existir à direita com suas próprias colunas "%". Usar a primeira
    # garante que pegamos a coluna correta (login_col + 2) e não a tabela dos dias.
    header_row = df_raw.iloc[header_row_idx]
    pct_col = None
    for col_idx, val in header_row.items():
        if col_idx <= login_col:
            continue
        if "%" in str(val):
            pct_col = col_idx
            break  # primeira coluna % após o login → tabela correta

    if pct_col is None:
        pct_col = login_col + 2  # fallback

    rows = []
    skip_tokens = {"nan", "Total Geral", "COP REDE RJ", "", "Rótulos de Linha"}
    for i in range(header_row_idx + 1, len(df_raw)):
        row = df_raw.iloc[i]
        login = str(row.get(login_col, "")).strip().upper()
        if not login or login in skip_tokens:
            continue
        pct_f = _parse_pct_value(row.get(pct_col, None))
        if pct_f is not None:
            rows.append({"Login": login, "DPA_Pct_Oficial": round(pct_f * 100, 2)})

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.merge(
        BASE_EQUIPE[["Matricula", "Nome", "Setor"]],
        left_on="Login", right_on="Matricula", how="left",
    ).drop(columns="Matricula")
    df["Nome"] = df["Nome"].fillna(
        df["Login"].str.upper().map(COORD_ANALYSTS_NAMES)
    ).fillna(df["Login"])
    df["Setor"] = df["Setor"].fillna("")
    df = df.sort_values("DPA_Pct_Oficial", ascending=False).reset_index(drop=True)
    df.index += 1
    df.index.name = "#"
    return df


def _parse_dpa_pivot_cache(raw_bytes: bytes) -> pd.DataFrame:
    """
    Lê o pivotCacheRecords2.xml do arquivo XLSX e retorna DataFrame com todos
    os campos brutos do cache de DPA (USUARIO_LOGIN, ANOMES, MESNOME,
    TEMPO_USO_SEC, HORARIO_JORNADA_SEC, DIA, etc.).

    Usa o cache 2 (pivotCacheDefinition2 / pivotCacheRecords2), que corresponde
    ao pivô de Ocupação DPA por Analista.
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import io
    import re

    ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    required_fields = {'USUARIO_LOGIN', 'TEMPO_USO_SEC', 'HORARIO_JORNADA_SEC', 'ANOMES'}

    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        # Tenta encontrar o pivot cache correto (pode ser cache 1, 2, 3...)
        cache_files = [n for n in zf.namelist() if re.match(r'xl/pivotCache/pivotCacheDefinition\d+\.xml', n)]
        def_xml = None
        rec_xml = None
        for cache_def in sorted(cache_files):
            cache_num = re.search(r'(\d+)', cache_def.split('/')[-1]).group(1)
            cache_rec = f'xl/pivotCache/pivotCacheRecords{cache_num}.xml'
            if cache_rec not in zf.namelist():
                continue
            test_def = zf.read(cache_def)
            test_tree = ET.fromstring(test_def)
            test_fields = {f.get('name') for f in test_tree.findall('.//x:cacheField', ns)}
            if required_fields.issubset(test_fields):
                def_xml = test_def
                rec_xml = zf.read(cache_rec)
                break

        if def_xml is None or rec_xml is None:
            return pd.DataFrame()

    tree_def = ET.fromstring(def_xml)
    fields   = tree_def.findall('.//x:cacheField', ns)
    field_names = [f.get('name') for f in fields]

    shared = {}
    for i, field in enumerate(fields):
        items = []
        si = field.find('x:sharedItems', ns)
        if si is not None:
            for child in si:
                tag = child.tag.split('}')[-1]
                if tag == 's':
                    items.append(child.get('v'))
                elif tag == 'n':
                    items.append(float(child.get('v')))
                elif tag in ('b', 'd'):
                    items.append(child.get('v'))
                else:
                    items.append(None)
        shared[i] = items

    root_rec = ET.fromstring(rec_xml)
    records = []
    for record in root_rec.findall('x:r', ns):
        row = {}
        for fi, child in enumerate(list(record)):
            if fi >= len(field_names):
                break
            fname = field_names[fi]
            tag = child.tag.split('}')[-1]
            if tag == 'x':
                ref = int(child.get('v', 0))
                row[fname] = shared[fi][ref] if ref < len(shared[fi]) else None
            elif tag == 'n':
                row[fname] = float(child.get('v', 0))
            elif tag == 's':
                row[fname] = child.get('v')
            elif tag == 'm':
                row[fname] = None
            else:
                row[fname] = child.get('v')
        records.append(row)

    df = pd.DataFrame(records)
    for c in ['ANOMES', 'ANO', 'MES', 'DIA',
              'TEMPO_USO_SEC', 'HORARIO_JORNADA_SEC', 'DIA_TRABALHADO']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    return df


def load_dpa_ocupacao(uploaded_file) -> tuple[pd.DataFrame, dict]:
    """
    Carrega o DPA por analista lendo DIRETAMENTE do pivot cache XML interno
    da planilha, ignorando o estado dos slicers/filtros do Excel.

    Detecta automaticamente o mês mais recente com dados reais (ANOMES máximo
    onde TEMPO_USO_SEC > 0), independente de como a planilha foi salva.

    Retorna:
    - df_analistas: DataFrame com Login, Nome, Setor, DPA_Pct_Oficial
    - mes_info: dict com mes_nome, mes_num, dpa_geral_pct
    """
    if hasattr(uploaded_file, 'read'):
        raw_bytes = uploaded_file.read()
    elif hasattr(uploaded_file, 'getvalue'):
        raw_bytes = uploaded_file.getvalue()
    else:
        raw_bytes = bytes(uploaded_file)

    df = _parse_dpa_pivot_cache(raw_bytes)

    # Verifica colunas essenciais do pivot cache
    _pivot_ok = not df.empty and all(
        c in df.columns for c in ['TEMPO_USO_SEC', 'HORARIO_JORNADA_SEC', 'ANOMES', 'USUARIO_LOGIN']
    )

    if not _pivot_ok:
        # Fallback: lê diretamente das abas "Consolidado" e "Analistas"
        try:
            df_cons = pd.read_excel(io.BytesIO(raw_bytes),
                                    sheet_name=DPA_SHEET_CONSOLIDADO, header=None)
            mes_info = _dpa_detect_mes_recente(df_cons)
        except Exception:
            mes_info = {"mes_nome": None, "mes_num": None, "dpa_geral_pct": None}
        try:
            df_anal = pd.read_excel(io.BytesIO(raw_bytes),
                                    sheet_name=DPA_SHEET_ANALISTAS, header=None)
            grp = _dpa_extract_analistas(df_anal)
        except Exception:
            grp = pd.DataFrame()
        return grp, mes_info

    # Mês mais recente com dados reais (descarta linhas sem uso registrado)
    df_real = df[df['TEMPO_USO_SEC'] > 0]
    if df_real.empty:
        return pd.DataFrame(), {"mes_nome": None, "mes_num": None, "dpa_geral_pct": None}

    anomes_max = df_real['ANOMES'].max()
    mes_num    = int(anomes_max) % 100
    mes_nome   = DPA_MESES_PT[mes_num - 1]

    df_mes = df[df['ANOMES'] == anomes_max].copy()

    # DPA geral da equipe toda neste mês
    total_uso     = df_mes['TEMPO_USO_SEC'].sum()
    total_jornada = df_mes['HORARIO_JORNADA_SEC'].sum()
    dpa_geral_pct = round(total_uso / total_jornada * 100, 2) if total_jornada > 0 else None

    mes_info = {
        "mes_nome":      mes_nome,
        "mes_num":       mes_num,
        "dpa_geral_pct": dpa_geral_pct,
    }

    # Normaliza login para maiúsculas antes do agrupamento/filtro (pivot cache pode vir em lowercase)
    df_mes['USUARIO_LOGIN'] = df_mes['USUARIO_LOGIN'].astype(str).str.strip().str.upper()

    # DPA por analista: soma tempo / soma jornada no mês
    grp = df_mes.groupby('USUARIO_LOGIN').agg(
        tempo_total=('TEMPO_USO_SEC', 'sum'),
        jornada_total=('HORARIO_JORNADA_SEC', 'sum'),
    ).reset_index()
    grp['DPA_Pct_Oficial'] = (
        grp['tempo_total'] / grp['jornada_total'] * 100
    ).where(grp['jornada_total'] > 0).round(2)

    # Filtra apenas analistas rastreados (equipe + coordenadores)
    grp = grp[grp['USUARIO_LOGIN'].isin(ALL_TRACKED_IDS)].copy()

    # Merge com nome e setor
    grp = grp.merge(
        BASE_EQUIPE[['Matricula', 'Nome', 'Setor']],
        left_on='USUARIO_LOGIN', right_on='Matricula', how='left',
    ).drop(columns='Matricula', errors='ignore')
    grp['Nome'] = grp['Nome'].fillna(
        grp['USUARIO_LOGIN'].str.upper().map(COORD_ANALYSTS_NAMES)
    ).fillna(grp['USUARIO_LOGIN'])
    grp['Setor'] = grp['Setor'].fillna(
        grp['USUARIO_LOGIN'].str.upper().map(_COORD_EMP_ANALYST_SECTOR)
    ).fillna('')
    grp = grp.rename(columns={'USUARIO_LOGIN': 'Login'})
    grp = grp.sort_values('DPA_Pct_Oficial', ascending=False).reset_index(drop=True)
    grp.index += 1
    grp.index.name = '#'

    return grp, mes_info


def dpa_ranking(df_analistas: pd.DataFrame) -> pd.DataFrame:
    """Retorna DataFrame de ranking de DPA já ordenado."""
    if df_analistas.empty:
        return pd.DataFrame()
    df = df_analistas.copy()
    df["Nome_Curto"] = df["Nome"].apply(primeiro_nome)
    return df[["Nome_Curto", "Login", "Setor", "DPA_Pct_Oficial"]].rename(
        columns={"Nome_Curto": "Analista", "DPA_Pct_Oficial": "DPA %"}
    )


def dpa_comparativo(df_analistas: pd.DataFrame, resumo_prod: pd.DataFrame) -> pd.DataFrame:
    """
    Junta DPA Oficial (planilha Ocupação) com DPA calculado (planilha Produtividade).
    Retorna DataFrame com ambos para comparação.
    """
    if df_analistas.empty or resumo_prod.empty:
        return pd.DataFrame()

    oficial = df_analistas[["Login", "Nome", "DPA_Pct_Oficial"]].copy()

    # Verifica se DPA_Media existe no resumo
    if "DPA_Media" not in resumo_prod.columns:
        return oficial

    calculado = resumo_prod[["USUARIO_LOGIN", "DPA_Media"]].rename(
        columns={"USUARIO_LOGIN": "Login", "DPA_Media": "DPA_Calculado"}
    )

    merged = oficial.merge(calculado, on="Login", how="left")
    merged["Diferença"] = (merged["DPA_Pct_Oficial"] - merged["DPA_Calculado"]).round(2)
    merged["Nome_Curto"] = merged["Nome"].apply(primeiro_nome)
    return merged[["Nome_Curto", "Login", "DPA_Pct_Oficial", "DPA_Calculado", "Diferença"]].rename(
        columns={"Nome_Curto": "Analista", "DPA_Pct_Oficial": "DPA Oficial %", "DPA_Calculado": "DPA Calculado %"}
    ).sort_values("DPA Oficial %", ascending=False).reset_index(drop=True)


# =====================================================
# Funções originais de produtividade
# =====================================================
def resumo_mensal(df: pd.DataFrame) -> pd.DataFrame:
    agg_dict = {COL_DATA: "count", COL_VOL_TOTAL: "sum"}
    for vc in VOL_COLS.keys():
        if vc in df.columns:
            agg_dict[vc] = "sum"
    group_cols = [COL_LOGIN, COL_NOME, "Setor", COL_MES, COL_ANOMES]
    existing_group = [c for c in group_cols if c in df.columns]
    g = df.groupby(existing_group).agg(agg_dict).reset_index()
    g = g.rename(columns={COL_DATA: "Dias_Trabalhados"})
    g["Media_Diaria"] = (g[COL_VOL_TOTAL] / g["Dias_Trabalhados"]).round(1)
    dpa_valid = df[(df[COL_DPA_RESULTADO] >= 0) & (df[COL_DPA_RESULTADO] <= 120)].copy()
    if not dpa_valid.empty:
        dpa_mean = dpa_valid.groupby(existing_group)[COL_DPA_RESULTADO].mean().reset_index()
        dpa_mean = dpa_mean.rename(columns={COL_DPA_RESULTADO: "DPA_Media"})
        dpa_mean["DPA_Media"] = dpa_mean["DPA_Media"].round(1)
        g = g.merge(dpa_mean, on=existing_group, how="left")
    else:
        g["DPA_Media"] = np.nan
    return g


def resumo_geral(df: pd.DataFrame) -> pd.DataFrame:
    agg_dict = {COL_DATA: "count", COL_VOL_TOTAL: "sum"}
    for vc in VOL_COLS.keys():
        if vc in df.columns:
            agg_dict[vc] = "sum"
    group_cols = [COL_LOGIN, COL_NOME, "Setor"]
    existing_group = [c for c in group_cols if c in df.columns]
    g = df.groupby(existing_group).agg(agg_dict).reset_index()
    g = g.rename(columns={COL_DATA: "Dias_Trabalhados"})
    g["Media_Diaria"] = (g[COL_VOL_TOTAL] / g["Dias_Trabalhados"]).round(1)
    dpa_valid = df[(df[COL_DPA_RESULTADO] >= 0) & (df[COL_DPA_RESULTADO] <= 120)].copy()
    if not dpa_valid.empty:
        dpa_mean = dpa_valid.groupby(existing_group)[COL_DPA_RESULTADO].mean().reset_index()
        dpa_mean = dpa_mean.rename(columns={COL_DPA_RESULTADO: "DPA_Media"})
        dpa_mean["DPA_Media"] = dpa_mean["DPA_Media"].round(1)
        g = g.merge(dpa_mean, on=existing_group, how="left")
    else:
        g["DPA_Media"] = np.nan
    return g


def evolucao_diaria(df: pd.DataFrame) -> pd.DataFrame:
    if COL_DATA not in df.columns:
        return pd.DataFrame()
    daily = df.groupby(COL_DATA).agg(
        Vol_Total=(COL_VOL_TOTAL, "sum"),
        Analistas=(COL_LOGIN, "nunique"),
    ).reset_index()
    daily["Media_por_Analista"] = (daily["Vol_Total"] / daily["Analistas"]).round(1)
    return daily


def composicao_volume(df: pd.DataFrame) -> pd.DataFrame:
    vol_data = {}
    for col, label in VOL_COLS.items():
        if col in df.columns:
            total = df[col].sum()
            if total > 0:
                vol_data[label] = total
    return pd.DataFrame(list(vol_data.items()), columns=["Atividade", "Volume"]).sort_values(
        "Volume", ascending=False
    )


def primeiro_nome(nome_completo: str) -> str:
    s = str(nome_completo).strip()
    parts = s.split()
    if len(parts) <= 2:
        return s
    return f"{parts[0]} {parts[-1]}"


# =====================================================
# INDICADORES TOA — Loader e processadores
# =====================================================

def _read_toa_sheet(uploaded_file) -> pd.DataFrame:
    """Lê a aba TOA tentando nomes candidatos; levanta ValueError se nenhum funcionar."""
    from src.config import TOA_SHEET_CANDIDATES
    import openpyxl, io
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    data = uploaded_file.read() if hasattr(uploaded_file, "read") else uploaded_file
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    available = wb.sheetnames
    wb.close()
    for candidate in TOA_SHEET_CANDIDATES:
        if candidate in available:
            return pd.read_excel(io.BytesIO(data), sheet_name=candidate)
    # Fallback: primeira aba
    return pd.read_excel(io.BytesIO(data), sheet_name=0)


def load_toa_indicadores(uploaded_file) -> pd.DataFrame:
    """
    Lê a planilha Analitico_Indicadores_TOA e retorna apenas
    TAREFAS CANCELADAS e TEMPO DE VALIDAÇÃO DO FORMULÁRIO
    filtrados pela equipe monitorada.

    Detecção automática do mês mais recente via coluna ANOMES.
    """
    df = _read_toa_sheet(uploaded_file)

    # Filtrar indicadores de interesse
    if TOA_COL_INDICADOR_NOME not in df.columns:
        return pd.DataFrame()

    df = df[df[TOA_COL_INDICADOR_NOME].isin(TOA_INDICADORES_FILTRO)].copy()
    if df.empty:
        return df

    # Normalizar login (maiúsculo, sem espaços)
    df[TOA_COL_LOGIN] = df[TOA_COL_LOGIN].astype(str).str.strip().str.upper()

    # Detectar ANOMES mais recente e filtrar
    if TOA_COL_ANOMES in df.columns:
        df[TOA_COL_ANOMES] = pd.to_numeric(df[TOA_COL_ANOMES], errors="coerce")
        anomes_recente = df[TOA_COL_ANOMES].max()
        df = df[df[TOA_COL_ANOMES] == anomes_recente].copy()

    # Filtrar equipe
    df = df[df[TOA_COL_LOGIN].isin(ALL_TRACKED_IDS)].copy()
    if df.empty:
        return df

    # Filtrar regional Leste
    if TOA_COL_REGIONAL in df.columns:
        df = df[df[TOA_COL_REGIONAL] == REGIONAL_FILTRO].copy()
    if df.empty:
        return df

    # Merge com nome e setor
    df = df.merge(
        BASE_EQUIPE[["Matricula", "Nome", "Setor"]],
        left_on=TOA_COL_LOGIN, right_on="Matricula", how="left"
    )
    # Para analistas fora da BASE_EQUIPE (coords): usa lookup de nomes, depois o login
    df["Nome"] = df["Nome"].fillna(
        df[TOA_COL_LOGIN].str.upper().map(COORD_ANALYSTS_NAMES)
    ).fillna(df[TOA_COL_LOGIN])
    df["Setor"] = df["Setor"].fillna(
        df[TOA_COL_LOGIN].str.upper().map(_COORD_EMP_ANALYST_SECTOR)
    ).fillna("")

    # Tipos
    if TOA_COL_INDICADOR in df.columns:
        df[TOA_COL_INDICADOR] = pd.to_numeric(df[TOA_COL_INDICADOR], errors="coerce")

    for c in [TOA_COL_DATA, TOA_COL_DT_CANCELAMENTO, TOA_COL_DT_INICIO_FORM, TOA_COL_DT_FIM_FORM]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    # TMR: normalizar para minutos, suportando todos os formatos comuns:
    #   - timedelta nativo
    #   - string "HH:MM:SS" / "H:MM:SS"
    #   - float fração de dia (Excel: 0.00347 ≈ 5 min)
    #   - float em segundos (ex: 318.0)
    #   - float já em minutos (ex: 5.3)
    if TOA_COL_TMR in df.columns:
        col = df[TOA_COL_TMR]
        if pd.api.types.is_timedelta64_dtype(col):
            df["TMR_min"] = col.dt.total_seconds() / 60
        elif pd.api.types.is_numeric_dtype(col):
            col_f = col.astype(float)
            mean_v = col_f.dropna().mean()
            if pd.notna(mean_v):
                if mean_v < 1:
                    # Fração de dia (Excel) → converte para minutos
                    df["TMR_min"] = col_f * 24 * 60
                elif mean_v < 300:
                    # Já em minutos
                    df["TMR_min"] = col_f
                else:
                    # Em segundos → converte para minutos
                    df["TMR_min"] = col_f / 60
            else:
                df["TMR_min"] = pd.Series(dtype=float, index=df.index)
        else:
            # Tenta string "HH:MM:SS"
            col_td = pd.to_timedelta(col, errors="coerce")
            if col_td.notna().any():
                df["TMR_min"] = col_td.dt.total_seconds() / 60
            else:
                # Fallback: tenta numérico dentro da string
                col_num = pd.to_numeric(col, errors="coerce")
                mean_v = col_num.dropna().mean()
                if pd.notna(mean_v) and mean_v < 1:
                    df["TMR_min"] = col_num * 24 * 60
                elif pd.notna(mean_v) and mean_v >= 300:
                    df["TMR_min"] = col_num / 60
                else:
                    df["TMR_min"] = col_num

    # Coluna ADERENTE normalizada:
    # Canceladas: INDICADOR=1 → NÃO ADERENTE → invertemos
    # Validação:  INDICADOR=1 → ADERENTE
    df["ADERENTE"] = df.apply(
        lambda row: (
            (row[TOA_COL_INDICADOR] == 0)
            if row[TOA_COL_INDICADOR_NOME] in TOA_IND_INVERTIDOS
            else (row[TOA_COL_INDICADOR] == 1)
        ),
        axis=1,
    ).astype(int)

    # Data para evolução diária
    if TOA_COL_DATA in df.columns:
        df["DATA_DIA"] = df[TOA_COL_DATA].dt.normalize()

    return df


def toa_anomes_recente(df: pd.DataFrame) -> int | None:
    """Retorna o ANOMES mais recente presente no DataFrame."""
    if df.empty or TOA_COL_ANOMES not in df.columns:
        return None
    v = pd.to_numeric(df[TOA_COL_ANOMES], errors="coerce").max()
    return int(v) if pd.notna(v) else None


def toa_resumo_por_indicador(df: pd.DataFrame) -> pd.DataFrame:
    """KPI geral por indicador: total, aderentes, aderência%, TMR médio."""
    if df.empty:
        return pd.DataFrame()
    rows = []
    for ind in TOA_INDICADORES_FILTRO:
        sub = df[df[TOA_COL_INDICADOR_NOME] == ind]
        if sub.empty:
            continue
        total = len(sub)
        ader  = int(sub["ADERENTE"].sum())
        pct   = round(ader / total * 100, 1) if total > 0 else 0.0
        tmr_m = sub["TMR_min"].mean() if "TMR_min" in sub.columns else None
        rows.append({
            "Indicador": ind,
            "Total": total,
            "Aderentes": ader,
            "Aderencia_Pct": pct,
            "TMR_Medio_min": round(tmr_m, 2) if tmr_m is not None and pd.notna(tmr_m) else None,
        })
    return pd.DataFrame(rows)


# ---- TAREFAS CANCELADAS ----

def toa_canceladas_por_analista(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ranking de tarefas canceladas por analista.
    Canceladas = todas as linhas deste indicador (INDICADOR=1 sempre).
    Menor = melhor.
    """
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS].copy()
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby([TOA_COL_LOGIN, "Nome", "Setor"]).agg(
        Canceladas=(TOA_COL_INDICADOR, "count"),
        TMR_Medio_h=("TMR_min", lambda x: round(x.mean() / 60, 2) if x.notna().any() else None),
    ).reset_index().rename(columns={TOA_COL_LOGIN: "Login"})
    return g.sort_values("Canceladas", ascending=False).reset_index(drop=True)


def toa_canceladas_por_tipo(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown de tarefas canceladas por TIPO_ATIVIDADE."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS]
    if sub.empty or TOA_COL_TIPO_ATIVIDADE not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_TIPO_ATIVIDADE).size().reset_index(name="Canceladas")
    g.columns = ["Tipo Atividade", "Canceladas"]
    return g.sort_values("Canceladas", ascending=False).reset_index(drop=True)


def toa_canceladas_por_aging(df: pd.DataFrame) -> pd.DataFrame:
    """Distribuição de tarefas canceladas por faixa de AGING."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS]
    if sub.empty or TOA_COL_AGING not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_AGING).size().reset_index(name="Canceladas")
    g.columns = ["Aging", "Canceladas"]
    # Ordenar pela ordem definida em config
    order_map = {v: i for i, v in enumerate(TOA_AGING_ORDER)}
    g["_ord"] = g["Aging"].map(order_map).fillna(99)
    return g.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)


def toa_canceladas_por_rede(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown de tarefas canceladas por REDE."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS]
    if sub.empty or TOA_COL_REDE not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_REDE).size().reset_index(name="Canceladas")
    g.columns = ["Rede", "Canceladas"]
    return g.sort_values("Canceladas", ascending=False).reset_index(drop=True)


def toa_canceladas_por_regional(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown de tarefas canceladas por Regional."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS]
    if sub.empty or TOA_COL_REGIONAL not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_REGIONAL).size().reset_index(name="Canceladas")
    g.columns = ["Regional", "Canceladas"]
    return g.sort_values("Canceladas", ascending=False).reset_index(drop=True)


def toa_canceladas_evolucao(df: pd.DataFrame) -> pd.DataFrame:
    """Evolução diária de tarefas canceladas."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS]
    if sub.empty or "DATA_DIA" not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby("DATA_DIA").size().reset_index(name="Canceladas")
    g.columns = ["Data", "Canceladas"]
    return g.sort_values("Data").reset_index(drop=True)


# ---- TEMPO DE VALIDAÇÃO DO FORMULÁRIO ----

def toa_validacao_por_analista(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ranking de aderência ao tempo de validação do formulário por analista.
    Inclui total de formulários, aderentes, aderência% e TMR médio em minutos.
    """
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO].copy()
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby([TOA_COL_LOGIN, "Nome", "Setor"]).agg(
        Total=(TOA_COL_INDICADOR, "count"),
        Aderentes=("ADERENTE", "sum"),
        TMR_Medio_min=("TMR_min", "mean"),
    ).reset_index().rename(columns={TOA_COL_LOGIN: "Login"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Total"] * 100).round(1)
    g["TMR_Medio_min"]  = g["TMR_Medio_min"].round(2)
    return g.sort_values("Aderencia_Pct", ascending=False).reset_index(drop=True)


def toa_validacao_por_tipo(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown do tempo de validação por TIPO_ATIVIDADE."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO]
    if sub.empty or TOA_COL_TIPO_ATIVIDADE not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_TIPO_ATIVIDADE).agg(
        Total=(TOA_COL_INDICADOR, "count"),
        Aderentes=("ADERENTE", "sum"),
        TMR_Medio_min=("TMR_min", "mean"),
    ).reset_index().rename(columns={TOA_COL_TIPO_ATIVIDADE: "Tipo Atividade"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Total"] * 100).round(1)
    g["TMR_Medio_min"]  = g["TMR_Medio_min"].round(2)
    return g.sort_values("Total", ascending=False).reset_index(drop=True)


def toa_validacao_por_rede(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown do tempo de validação por REDE."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO]
    if sub.empty or TOA_COL_REDE not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_REDE).agg(
        Total=(TOA_COL_INDICADOR, "count"),
        Aderentes=("ADERENTE", "sum"),
        TMR_Medio_min=("TMR_min", "mean"),
    ).reset_index().rename(columns={TOA_COL_REDE: "Rede"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Total"] * 100).round(1)
    g["TMR_Medio_min"]  = g["TMR_Medio_min"].round(2)
    return g.sort_values("Total", ascending=False).reset_index(drop=True)


def toa_validacao_por_regional(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown do tempo de validação por Regional."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO]
    if sub.empty or TOA_COL_REGIONAL not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(TOA_COL_REGIONAL).agg(
        Total=(TOA_COL_INDICADOR, "count"),
        Aderentes=("ADERENTE", "sum"),
        TMR_Medio_min=("TMR_min", "mean"),
    ).reset_index().rename(columns={TOA_COL_REGIONAL: "Regional"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Total"] * 100).round(1)
    g["TMR_Medio_min"]  = g["TMR_Medio_min"].round(2)
    return g.sort_values("Total", ascending=False).reset_index(drop=True)


def toa_validacao_evolucao(df: pd.DataFrame) -> pd.DataFrame:
    """Evolução diária da aderência ao tempo de validação."""
    sub = df[df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO]
    if sub.empty or "DATA_DIA" not in sub.columns:
        return pd.DataFrame()
    sub = sub.dropna(subset=["DATA_DIA"])
    g = sub.groupby("DATA_DIA").agg(
        Total=(TOA_COL_INDICADOR, "count"),
        Aderentes=("ADERENTE", "sum"),
        TMR_Medio_min=("TMR_min", "mean"),
    ).reset_index().rename(columns={"DATA_DIA": "Data"})
    g["Aderencia_Pct"] = (g["Aderentes"] / g["Total"] * 100).round(1)
    g["TMR_Medio_min"]  = g["TMR_Medio_min"].round(2)
    return g.sort_values("Data").reset_index(drop=True)


# =====================================================
# FECHAMENTO TOA x SIR — Loader e processadores
# =====================================================

def _parse_pivot_cache(raw_bytes: bytes) -> pd.DataFrame:
    """
    Extrai registros brutos do pivot cache interno de um arquivo xlsx.
    Retorna DataFrame com todas as colunas presentes no cache.
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import io

    ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        # Ler definição do pivot cache
        def_xml = zf.read('xl/pivotCache/pivotCacheDefinition1.xml')
        rec_xml = zf.read('xl/pivotCache/pivotCacheRecords1.xml')

    tree_def = ET.fromstring(def_xml)
    fields   = tree_def.findall('.//x:cacheField', ns)
    field_names = [f.get('name') for f in fields]

    # Construir itens compartilhados por campo
    shared = {}
    for i, field in enumerate(fields):
        items = []
        si = field.find('x:sharedItems', ns)
        if si is not None:
            for child in si:
                tag = child.tag.split('}')[-1]
                if tag == 's':
                    items.append(child.get('v'))
                elif tag == 'n':
                    items.append(float(child.get('v')))
                elif tag in ('b', 'd'):
                    items.append(child.get('v'))
                else:
                    items.append(None)
        shared[i] = items

    # Parsear registros
    root_rec = ET.fromstring(rec_xml)
    records = []
    for record in root_rec.findall('x:r', ns):
        row = {}
        for fi, child in enumerate(list(record)):
            if fi >= len(field_names):
                break
            fname = field_names[fi]
            tag = child.tag.split('}')[-1]
            if tag == 'x':       # referência a item compartilhado
                ref = int(child.get('v', 0))
                row[fname] = shared[fi][ref] if ref < len(shared[fi]) else None
            elif tag == 'n':
                row[fname] = float(child.get('v', 0))
            elif tag == 's':
                row[fname] = child.get('v')
            elif tag == 'm':
                row[fname] = None
            else:
                row[fname] = child.get('v')
        records.append(row)

    return pd.DataFrame(records)


def load_fechamento_toa_sir(
    uploaded_file,
    team_ids: set | None = None,
    turnos: set | None = None,
) -> pd.DataFrame:
    """
    Carrega a planilha Fechamento_TOA_x_SIR.xlsx lendo o pivot cache interno.
    Filtra automaticamente:
      - TURNO = turnos (default: {'Madrugada'})
      - ANOMES = mais recente
      - LOGIN_VALIDOU_FECHAMENTO = team_ids (default: EQUIPE_IDS)
    Retorna DataFrame pronto para análise com coluna ASSERTIVO (0/1).
    """
    from src.config import (
        EQUIPE_IDS, BASE_EQUIPE,
        FECH_SIR_COL_LOGIN, FECH_SIR_COL_TURNO, FECH_SIR_COL_ANOMES,
        FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER,
        FECH_SIR_TURNO_MADRUGADA,
    )

    if hasattr(uploaded_file, 'read'):
        raw_bytes = uploaded_file.read()
    else:
        raw_bytes = uploaded_file

    df = _parse_pivot_cache(raw_bytes)
    if df.empty:
        return pd.DataFrame()

    # Normalizar login
    df[FECH_SIR_COL_LOGIN] = df[FECH_SIR_COL_LOGIN].astype(str).str.strip().str.upper()

    # Filtrar ANOMES mais recente
    if FECH_SIR_COL_ANOMES in df.columns:
        df[FECH_SIR_COL_ANOMES] = pd.to_numeric(df[FECH_SIR_COL_ANOMES], errors='coerce')
        anomes_max = df[FECH_SIR_COL_ANOMES].max()
        df = df[df[FECH_SIR_COL_ANOMES] == anomes_max].copy()

    # Filtrar turno — busca coluna TURNO ignorando maiúsculas/espaços
    _turno_col = next(
        (c for c in df.columns if c.strip().upper() == FECH_SIR_COL_TURNO.upper()), None
    )
    _turno_filter = turnos if turnos is not None else {FECH_SIR_TURNO_MADRUGADA}
    _turno_filter_upper = {str(t).strip().upper() for t in _turno_filter}
    if _turno_col:
        df = df[df[_turno_col].astype(str).str.strip().str.upper().isin(_turno_filter_upper)].copy()
    else:
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    # Filtrar equipe (case-insensitive)
    _team = team_ids if team_ids is not None else EQUIPE_IDS
    df = df[df[FECH_SIR_COL_LOGIN].str.upper().isin({e.upper() for e in _team})].copy()
    if df.empty:
        return pd.DataFrame()

    # Filtrar regional Leste
    from src.config import FECH_SIR_COL_REGIONAL, REGIONAL_FILTRO as _REGIONAL
    if FECH_SIR_COL_REGIONAL in df.columns:
        df = df[df[FECH_SIR_COL_REGIONAL] == _REGIONAL].copy()
    if df.empty:
        return pd.DataFrame()

    # Merge com nome e setor
    base = BASE_EQUIPE.copy()
    base['Matricula_upper'] = base['Matricula'].str.upper()
    df['_login_up'] = df[FECH_SIR_COL_LOGIN].str.upper()
    df = df.merge(
        base[['Matricula_upper', 'Matricula', 'Nome', 'Setor']],
        left_on='_login_up', right_on='Matricula_upper', how='left'
    ).drop(columns=['_login_up', 'Matricula_upper'])
    df["Nome"] = df["Nome"].fillna(
        df[FECH_SIR_COL_LOGIN].str.upper().map(COORD_ANALYSTS_NAMES)
    ).fillna(df[FECH_SIR_COL_LOGIN])
    df["Setor"] = df["Setor"].fillna("")

    # Garantir tipos numéricos
    for c in [FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    # Coluna de volume assertivo — já é um número (soma de tarefas assertivas por registro)
    df['ASSERTIVO'] = df[FECH_SIR_COL_ASSERTIVO]
    if FECH_SIR_COL_NAO_ASSER in df.columns:
        df['NAO_ASSERTIVO'] = df[FECH_SIR_COL_NAO_ASSER]
    else:
        df['NAO_ASSERTIVO'] = (df[FECH_SIR_COL_VOLUME] - df['ASSERTIVO']).clip(lower=0)

    return df


def _fech_sir_nome_display(df: pd.DataFrame, login_col: str) -> "pd.Series":
    """
    Tenta encontrar uma coluna de nome legível no DataFrame do Fechamento SIR.
    Procura por qualquer coluna cujo nome contenha 'NOME' (case-insensitive),
    excluindo colunas puramente numéricas. Retorna a série de login como fallback.
    """
    metric_keywords = {'VOLUME', 'ASSERTIVO', 'NAO_ASSER', 'ANOMES', 'DIA', 'MES'}
    for col in df.columns:
        col_up = col.strip().upper()
        if 'NOME' in col_up and not any(kw in col_up for kw in metric_keywords):
            series = df[col].astype(str).str.strip()
            # Confirmar que não são puramente numéricos
            if not series.str.match(r'^\d+\.?\d*$').all():
                return series
    return df[login_col]


def load_fora_equipe_fech_sir(raw_bytes) -> pd.DataFrame:
    """
    Carrega a planilha Fechamento_TOA_x_SIR.xlsx e retorna apenas analistas
    que NÃO fazem parte da equipe (EQUIPE_IDS), filtrando pela madrugada.
    Útil para o admin visualizar quem de fora da equipe aparece na madrugada.
    """
    from src.config import (
        EQUIPE_IDS,
        FECH_SIR_COL_LOGIN, FECH_SIR_COL_TURNO, FECH_SIR_COL_ANOMES,
        FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER,
        FECH_SIR_TURNO_MADRUGADA, FECH_SIR_COL_REGIONAL, REGIONAL_FILTRO,
    )

    if hasattr(raw_bytes, 'read'):
        raw_bytes = raw_bytes.read()

    df = _parse_pivot_cache(raw_bytes)
    if df.empty:
        return pd.DataFrame()

    # Normalizar login
    df[FECH_SIR_COL_LOGIN] = df[FECH_SIR_COL_LOGIN].astype(str).str.strip().str.upper()

    # Filtrar ANOMES mais recente
    if FECH_SIR_COL_ANOMES in df.columns:
        df[FECH_SIR_COL_ANOMES] = pd.to_numeric(df[FECH_SIR_COL_ANOMES], errors='coerce')
        anomes_max = df[FECH_SIR_COL_ANOMES].max()
        df = df[df[FECH_SIR_COL_ANOMES] == anomes_max].copy()

    # Filtrar madrugada — busca coluna TURNO ignorando maiúsculas/espaços
    _turno_col2 = next(
        (c for c in df.columns if c.strip().upper() == FECH_SIR_COL_TURNO.upper()), None
    )
    if _turno_col2:
        df = df[
            df[_turno_col2].astype(str).str.strip().str.upper()
            == FECH_SIR_TURNO_MADRUGADA.upper()
        ].copy()
    else:
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    # Filtrar regional Leste (case-insensitive)
    if FECH_SIR_COL_REGIONAL in df.columns:
        df = df[
            df[FECH_SIR_COL_REGIONAL].astype(str).str.strip().str.upper()
            == REGIONAL_FILTRO.upper()
        ].copy()

    if df.empty:
        return pd.DataFrame()

    # Manter apenas quem NÃO é da equipe
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df = df[~df[FECH_SIR_COL_LOGIN].isin(equipe_upper)].copy()

    if df.empty:
        return pd.DataFrame()

    # Garantir tipos numéricos
    for c in [FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    df['ASSERTIVO'] = df[FECH_SIR_COL_ASSERTIVO] if FECH_SIR_COL_ASSERTIVO in df.columns else 0
    if FECH_SIR_COL_NAO_ASSER in df.columns:
        df['NAO_ASSERTIVO'] = df[FECH_SIR_COL_NAO_ASSER]
    else:
        df['NAO_ASSERTIVO'] = (df[FECH_SIR_COL_VOLUME] - df['ASSERTIVO']).clip(lower=0)

    # Coluna de exibição: nome legível se disponível, senão o próprio login
    df['NOME_DISPLAY'] = _fech_sir_nome_display(df, FECH_SIR_COL_LOGIN)

    return df


def load_fora_equipe_fech_sir_coord(raw_bytes) -> pd.DataFrame:
    """
    Retorna registros de analistas externos (não da equipe) que aparecem na
    Madrugada (22:00–05:59), mas mostrando apenas seus casos nos turnos Manhã/Tarde
    (06:00–21:59). Usado para coordenadores verem o que esses analistas fazem
    no horário diurno.
    """
    from src.config import (
        EQUIPE_IDS,
        FECH_SIR_COL_LOGIN, FECH_SIR_COL_TURNO,
        FECH_SIR_COL_ANOMES, FECH_SIR_COL_VOLUME,
        FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER,
        FECH_SIR_TURNO_MADRUGADA, FECH_SIR_COL_REGIONAL, REGIONAL_FILTRO,
    )

    if hasattr(raw_bytes, 'read'):
        raw_bytes = raw_bytes.read()

    df = _parse_pivot_cache(raw_bytes)
    if df.empty:
        return pd.DataFrame()

    # Normalizar login
    df[FECH_SIR_COL_LOGIN] = df[FECH_SIR_COL_LOGIN].astype(str).str.strip().str.upper()

    # Filtrar ANOMES mais recente
    if FECH_SIR_COL_ANOMES in df.columns:
        df[FECH_SIR_COL_ANOMES] = pd.to_numeric(df[FECH_SIR_COL_ANOMES], errors='coerce')
        anomes_max = df[FECH_SIR_COL_ANOMES].max()
        df = df[df[FECH_SIR_COL_ANOMES] == anomes_max].copy()

    # Filtrar regional Leste (case-insensitive)
    if FECH_SIR_COL_REGIONAL in df.columns:
        df = df[
            df[FECH_SIR_COL_REGIONAL].astype(str).str.strip().str.upper()
            == REGIONAL_FILTRO.upper()
        ].copy()

    if df.empty:
        return pd.DataFrame()

    # Excluir membros da equipe
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df = df[~df[FECH_SIR_COL_LOGIN].isin(equipe_upper)].copy()

    if df.empty:
        return pd.DataFrame()

    # Encontrar coluna TURNO
    _turno_col = next(
        (c for c in df.columns if c.strip().upper() == FECH_SIR_COL_TURNO.upper()), None
    )
    if not _turno_col:
        return pd.DataFrame()

    df[_turno_col] = df[_turno_col].astype(str).str.strip()

    # Identificar logins que aparecem na Madrugada
    logins_madrugada = set(
        df.loc[df[_turno_col].str.upper() == FECH_SIR_TURNO_MADRUGADA.upper(), FECH_SIR_COL_LOGIN]
    )
    if not logins_madrugada:
        return pd.DataFrame()

    # Retornar registros desses logins mas fora do turno Madrugada (Manhã + Tarde)
    df_coord = df[
        df[FECH_SIR_COL_LOGIN].isin(logins_madrugada) &
        (df[_turno_col].str.upper() != FECH_SIR_TURNO_MADRUGADA.upper())
    ].copy()

    if df_coord.empty:
        return pd.DataFrame()

    # Garantir tipos numéricos
    for c in [FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO, FECH_SIR_COL_NAO_ASSER]:
        if c in df_coord.columns:
            df_coord[c] = pd.to_numeric(df_coord[c], errors='coerce').fillna(0)

    df_coord['ASSERTIVO'] = df_coord[FECH_SIR_COL_ASSERTIVO] if FECH_SIR_COL_ASSERTIVO in df_coord.columns else 0
    if FECH_SIR_COL_NAO_ASSER in df_coord.columns:
        df_coord['NAO_ASSERTIVO'] = df_coord[FECH_SIR_COL_NAO_ASSER]
    else:
        df_coord['NAO_ASSERTIVO'] = (df_coord[FECH_SIR_COL_VOLUME] - df_coord['ASSERTIVO']).clip(lower=0)

    # Coluna de exibição: nome legível se disponível, senão o próprio login
    df_coord['NOME_DISPLAY'] = _fech_sir_nome_display(df_coord, FECH_SIR_COL_LOGIN)

    return df_coord


def fora_equipe_resumo_por_login(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa dados de analistas externos (fora da equipe) por nome/login,
    retornando volume total, ganhos (assertivos) e perdas (não assertivos).
    Usa NOME_DISPLAY se disponível (nome legível), senão usa o login.
    """
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_LOGIN, FECH_SIR_COL_VOLUME
    group_col = 'NOME_DISPLAY' if 'NOME_DISPLAY' in df.columns else FECH_SIR_COL_LOGIN
    g = df.groupby(group_col).agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Ganhos=('ASSERTIVO', 'sum'),
        Perdas=('NAO_ASSERTIVO', 'sum'),
    ).reset_index()
    g.columns = ['Login', 'Volume', 'Ganhos', 'Perdas']
    g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


# =====================================================
# ANALISTAS EXTERNOS — ETIT POR EVENTO
# =====================================================
def load_fora_equipe_etit(uploaded_file) -> pd.DataFrame:
    """ETIT POR EVENTO para analistas fora da equipe (admin only)."""
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    sheets = list_sheets(uploaded_file)
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    sheet_to_read = None
    for candidate in ETIT_SHEET_CANDIDATES:
        if candidate in sheets:
            sheet_to_read = candidate
            break
    if sheet_to_read is None:
        sheet_to_read = sheets[0]

    df = pd.read_excel(uploaded_file, sheet_name=sheet_to_read)

    if ETIT_COL_INDICADOR not in df.columns:
        return pd.DataFrame()
    df = df[df[ETIT_COL_INDICADOR] == ETIT_INDICADOR_FILTRO].copy()
    if df.empty:
        return pd.DataFrame()

    # Normalizar login
    df[ETIT_COL_LOGIN] = (
        df[ETIT_COL_LOGIN].astype(str).str.strip()
        .map(lambda x, a=LOGIN_ALIASES: a.get(x, x))
    )

    # Filtrar regional Leste
    if ETIT_COL_REGIONAL in df.columns:
        df = df[df[ETIT_COL_REGIONAL] == REGIONAL_FILTRO].copy()
    if df.empty:
        return pd.DataFrame()

    # Filtro madrugada (22:00–05:59)
    from src.config import FECH_SIR_TURNO_MADRUGADA
    if ETIT_COL_TURNO in df.columns:
        df = df[df[ETIT_COL_TURNO].astype(str).str.strip().str.upper() == FECH_SIR_TURNO_MADRUGADA.upper()].copy()
    elif ETIT_COL_DT_INICIO in df.columns:
        _dt = pd.to_datetime(df[ETIT_COL_DT_INICIO], errors='coerce')
        _h = _dt.dt.hour
        df = df[_dt.notna() & ((_h >= 22) | (_h <= 5))].copy()
    if df.empty:
        return pd.DataFrame()

    # Manter apenas quem NÃO é da equipe
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df = df[~df[ETIT_COL_LOGIN].str.upper().isin(equipe_upper)].copy()
    if df.empty:
        return pd.DataFrame()

    for c in [ETIT_COL_TMA, ETIT_COL_TMR, ETIT_COL_VOLUME, ETIT_COL_INDICADOR_VAL]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    for c in [ETIT_COL_DT_INICIO, ETIT_COL_DT_FIM, ETIT_COL_DT_ACIONAMENTO]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors='coerce')
    if ETIT_COL_ANOMES in df.columns:
        df[ETIT_COL_ANOMES] = (
            df[ETIT_COL_ANOMES]
            .astype(str).str.strip()
            .str.replace(r"\.0+$", "", regex=True)
        )
    return df


def fora_equipe_resumo_etit(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega ETIT externo por login: Volume, Aderentes (Ganhos), Não Aderentes (Perdas)."""
    if df.empty:
        return pd.DataFrame()
    g = df.groupby(ETIT_COL_LOGIN).agg(
        Volume=(ETIT_COL_VOLUME, 'sum'),
        Ganhos=(ETIT_COL_INDICADOR_VAL, 'sum'),
    ).reset_index()
    g.columns = ['Login', 'Volume', 'Ganhos']
    g['Volume'] = g['Volume'].fillna(0).astype(int)
    g['Ganhos'] = g['Ganhos'].fillna(0).astype(int)
    g['Perdas'] = (g['Volume'] - g['Ganhos']).clip(lower=0)
    g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


# =====================================================
# ANALISTAS EXTERNOS — INDICADORES RESIDENCIAL
# (df_res_ind já carrega todos — só filtrar fora da equipe)
# =====================================================
def fora_equipe_resumo_res(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recebe o df completo de Indicadores Residencial (sem filtro de equipe)
    e retorna resumo por ID_MOSTRA para quem NÃO é da equipe.
    """
    if df.empty or RES_COL_ID_MOSTRA not in df.columns:
        return pd.DataFrame()
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df_ext = df[~df[RES_COL_ID_MOSTRA].astype(str).str.upper().isin(equipe_upper)].copy()
    if df_ext.empty:
        return pd.DataFrame()
    g = df_ext.groupby(RES_COL_ID_MOSTRA).agg(
        Volume=(RES_COL_VOLUME, 'sum'),
        Ganhos=('ADERENTE', 'sum'),
    ).reset_index()
    g.columns = ['Login', 'Volume', 'Ganhos']
    g['Volume'] = g['Volume'].fillna(0).astype(int)
    g['Ganhos'] = g['Ganhos'].fillna(0).astype(int)
    g['Perdas'] = (g['Volume'] - g['Ganhos']).clip(lower=0)
    g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


def _res_agg_por_indicador(df_ext: pd.DataFrame) -> dict:
    """Agrega Volume/Ganhos/Perdas/Assertividade por indicador, agrupando por RES_COL_LOGIN."""
    result = {}
    for ind in RES_INDICADORES_FILTRO:
        df_ind = df_ext[df_ext[RES_COL_INDICADOR_NOME] == ind]
        if df_ind.empty:
            continue
        g = df_ind.groupby(RES_COL_LOGIN).agg(
            Volume=(RES_COL_VOLUME, 'sum'),
            Ganhos=('ADERENTE', 'sum'),
        ).reset_index()
        g.columns = ['Login', 'Volume', 'Ganhos']
        g['Volume'] = g['Volume'].fillna(0).astype(int)
        g['Ganhos'] = g['Ganhos'].fillna(0).astype(int)
        g['Perdas'] = (g['Volume'] - g['Ganhos']).clip(lower=0)
        g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
        result[ind] = g.sort_values('Volume', ascending=False).reset_index(drop=True)
    return result


def fora_equipe_resumo_res_por_indicador_adm(df: pd.DataFrame) -> dict:
    """
    Analistas externos (Madrugada 22:00–05:59) nos Indicadores Residencial.
    Usa RES_COL_LOGIN (LOGIN_PRIMEIRO_ACIONAMENTO_FO para HFC,
    LOGIN_PRIMEIRO_ACIONAMENTO_GPON para GPON/ASSERTIVIDADE).
    Retorna dict {indicador: resumo_df} com colunas: Login, Volume, Ganhos, Perdas, Assertividade_Pct.
    """
    if df.empty or RES_COL_LOGIN not in df.columns or RES_COL_TURNO not in df.columns:
        return {}
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df_ext = df[
        (~df[RES_COL_LOGIN].isin(equipe_upper))
        & (df[RES_COL_LOGIN] != "")
        & (df[RES_COL_TURNO] == "Madrugada")
    ].copy()
    if df_ext.empty:
        return {}
    return _res_agg_por_indicador(df_ext)


def fora_equipe_resumo_res_por_indicador_coord(df: pd.DataFrame) -> dict:
    """
    Analistas externos que aparecem na Madrugada, mostrando sua atividade
    diurna (06:00–21:59) nos Indicadores Residencial.
    Usa RES_COL_LOGIN (LOGIN_PRIMEIRO_ACIONAMENTO_FO para HFC,
    LOGIN_PRIMEIRO_ACIONAMENTO_GPON para GPON/ASSERTIVIDADE).
    Retorna dict {indicador: resumo_df} com colunas: Login, Volume, Ganhos, Perdas, Assertividade_Pct.
    """
    if df.empty or RES_COL_LOGIN not in df.columns or RES_COL_TURNO not in df.columns:
        return {}
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df_ext = df[
        (~df[RES_COL_LOGIN].isin(equipe_upper))
        & (df[RES_COL_LOGIN] != "")
    ].copy()
    if df_ext.empty:
        return {}
    # Identificar logins que aparecem na Madrugada
    logins_madrugada = set(
        df_ext.loc[df_ext[RES_COL_TURNO] == "Madrugada", RES_COL_LOGIN]
    )
    if not logins_madrugada:
        return {}
    # Retornar registros desses logins nos turnos Manhã/Tarde (diurno)
    df_coord = df_ext[
        df_ext[RES_COL_LOGIN].isin(logins_madrugada)
        & (df_ext[RES_COL_TURNO] != "Madrugada")
    ].copy()
    if df_coord.empty:
        return {}
    return _res_agg_por_indicador(df_coord)


# =====================================================
# ANALISTAS EXTERNOS — INDICADORES TOA
# =====================================================
def load_fora_equipe_toa(uploaded_file) -> pd.DataFrame:
    """TOA CANCELADAS / VALIDAÇÃO para analistas fora da equipe (admin only)."""
    df = _read_toa_sheet(uploaded_file)

    if TOA_COL_INDICADOR_NOME not in df.columns:
        return pd.DataFrame()
    df = df[df[TOA_COL_INDICADOR_NOME].isin(TOA_INDICADORES_FILTRO)].copy()
    if df.empty:
        return df

    df[TOA_COL_LOGIN] = df[TOA_COL_LOGIN].astype(str).str.strip().str.upper()

    # ANOMES mais recente
    if TOA_COL_ANOMES in df.columns:
        df[TOA_COL_ANOMES] = pd.to_numeric(df[TOA_COL_ANOMES], errors='coerce')
        anomes_recente = df[TOA_COL_ANOMES].max()
        df = df[df[TOA_COL_ANOMES] == anomes_recente].copy()

    # Regional Leste
    if TOA_COL_REGIONAL in df.columns:
        df = df[df[TOA_COL_REGIONAL] == REGIONAL_FILTRO].copy()
    if df.empty:
        return df

    # Manter apenas quem NÃO é da equipe
    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    df = df[~df[TOA_COL_LOGIN].isin(equipe_upper)].copy()
    if df.empty:
        return df

    if TOA_COL_INDICADOR in df.columns:
        df[TOA_COL_INDICADOR] = pd.to_numeric(df[TOA_COL_INDICADOR], errors='coerce')

    for c in [TOA_COL_DATA, TOA_COL_DT_CANCELAMENTO, TOA_COL_DT_INICIO_FORM, TOA_COL_DT_FIM_FORM]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors='coerce')

    # Filtro madrugada (22:00–05:59) por indicador
    from src.config import FECH_SIR_TURNO_MADRUGADA
    _turno_col = next((c for c in df.columns if c.upper() == 'TURNO'), None)
    if _turno_col:
        # Arquivo TOA possui coluna TURNO — usar diretamente (igual a ETIT/Fechamento)
        df = df[df[_turno_col].astype(str).str.strip().str.upper() == FECH_SIR_TURNO_MADRUGADA.upper()].copy()
    else:
        # Fallback: filtrar pelo horário das colunas de data específicas por indicador
        def _is_mad_h(s: pd.Series) -> pd.Series:
            h = s.dt.hour
            # Exclui NaT e registros com hora fora do intervalo 22:00–05:59
            return s.notna() & ((h >= 22) | (h <= 5))

        canc_mask  = df[TOA_COL_INDICADOR_NOME] == TOA_IND_CANCELADAS
        valid_mask = df[TOA_COL_INDICADOR_NOME] == TOA_IND_VALIDACAO
        canc_time_ok  = _is_mad_h(df[TOA_COL_DT_CANCELAMENTO]) if TOA_COL_DT_CANCELAMENTO in df.columns else pd.Series(False, index=df.index)
        valid_time_ok = _is_mad_h(df[TOA_COL_DT_INICIO_FORM])  if TOA_COL_DT_INICIO_FORM  in df.columns else pd.Series(False, index=df.index)
        df = df[(canc_mask & canc_time_ok) | (valid_mask & valid_time_ok)].copy()
    if df.empty:
        return df

    df['ADERENTE'] = df.apply(
        lambda row: (
            (row[TOA_COL_INDICADOR] == 0)
            if row[TOA_COL_INDICADOR_NOME] in TOA_IND_INVERTIDOS
            else (row[TOA_COL_INDICADOR] == 1)
        ),
        axis=1,
    ).astype(int)
    return df


def fora_equipe_resumo_toa(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega TOA externo por login: Total (Volume), Aderentes (Ganhos), Não Aderentes (Perdas)."""
    if df.empty:
        return pd.DataFrame()
    g = df.groupby(TOA_COL_LOGIN).agg(
        Volume=(TOA_COL_LOGIN, 'count'),
        Ganhos=('ADERENTE', 'sum'),
    ).reset_index()
    g.columns = ['Login', 'Volume', 'Ganhos']
    g['Perdas'] = (g['Volume'] - g['Ganhos']).clip(lower=0)
    g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


def fora_equipe_resumo_toa_por_indicador(df: pd.DataFrame) -> dict:
    """
    Retorna dict {indicador: resumo_df} para analistas externos nos Indicadores TOA.
    Cada resumo_df tem colunas: Login, Volume, Ganhos, Perdas, Assertividade_Pct.
    """
    if df.empty:
        return {}
    result = {}
    for ind in TOA_INDICADORES_FILTRO:
        df_ind = df[df[TOA_COL_INDICADOR_NOME] == ind]
        if df_ind.empty:
            continue
        g = df_ind.groupby(TOA_COL_LOGIN).agg(
            Volume=(TOA_COL_LOGIN, 'count'),
            Ganhos=('ADERENTE', 'sum'),
        ).reset_index()
        g.columns = ['Login', 'Volume', 'Ganhos']
        g['Perdas'] = (g['Volume'] - g['Ganhos']).clip(lower=0)
        g['Assertividade_Pct'] = (g['Ganhos'] / g['Volume'] * 100).where(g['Volume'] > 0, 0).round(1)
        result[ind] = g.sort_values('Volume', ascending=False).reset_index(drop=True)
    return result


# =====================================================
# ANALISTAS EXTERNOS — OCUPAÇÃO DPA
# =====================================================
def load_fora_equipe_dpa(uploaded_file) -> pd.DataFrame:
    """
    Lê a aba Analistas da planilha DPA e retorna login + DPA%
    apenas para analistas que NÃO são da equipe.
    """
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    df_raw = pd.read_excel(uploaded_file, sheet_name=DPA_SHEET_ANALISTAS, header=None)

    header_row_idx = None
    login_col = None
    for i, row in df_raw.iterrows():
        for col_idx, val in row.items():
            if str(val).strip() == 'Rótulos de Linha':
                header_row_idx = i
                login_col = col_idx
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        return pd.DataFrame()

    header_row = df_raw.iloc[header_row_idx]
    pct_col = None
    for col_idx, val in header_row.items():
        if col_idx <= login_col:
            continue
        if '%' in str(val):
            pct_col = col_idx
            break
    if pct_col is None:
        pct_col = login_col + 2

    equipe_upper = {e.upper() for e in EQUIPE_IDS}
    skip_tokens = {'nan', 'Total Geral', 'COP REDE RJ', '', 'Rótulos de Linha'}
    rows = []
    for i in range(header_row_idx + 1, len(df_raw)):
        row = df_raw.iloc[i]
        login = str(row.get(login_col, '')).strip().upper()
        if not login or login in skip_tokens or login in equipe_upper:
            continue
        pct_f = _parse_pct_value(row.get(pct_col, None))
        if pct_f is not None:
            rows.append({'Login': login, 'DPA_Pct': round(pct_f * 100, 2)})
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values('DPA_Pct', ascending=False).reset_index(drop=True)


def fech_sir_resumo_analista(df: pd.DataFrame) -> pd.DataFrame:
    """Resumo de assertividade por analista."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_LOGIN, FECH_SIR_COL_VOLUME, FECH_SIR_COL_ASSERTIVO
    g = df.groupby([FECH_SIR_COL_LOGIN, 'Nome', 'Setor']).agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Assertivos=('ASSERTIVO', 'sum'),
    ).reset_index().rename(columns={FECH_SIR_COL_LOGIN: 'Login'})
    g['Assertividade_Pct'] = (g['Assertivos'] / g['Volume'] * 100).round(1)
    return g.sort_values('Assertividade_Pct', ascending=False).reset_index(drop=True)


def fech_sir_por_causa_toa(df: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    """Volume não assertivo por causa TOA."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_CAUSA_TOA
    nao_asser = df[df['NAO_ASSERTIVO'] > 0] if 'NAO_ASSERTIVO' in df.columns else df[df['ASSERTIVO'] == 0]
    if nao_asser.empty or FECH_SIR_COL_CAUSA_TOA not in nao_asser.columns:
        return pd.DataFrame()
    vol_col = 'NAO_ASSERTIVO' if 'NAO_ASSERTIVO' in nao_asser.columns else 'ASSERTIVO'
    g = nao_asser.groupby(FECH_SIR_COL_CAUSA_TOA)[vol_col].sum().reset_index()
    g.columns = ['Causa TOA', 'Não Assertivo']
    g['Não Assertivo'] = g['Não Assertivo'].astype(int)
    return g.sort_values('Não Assertivo', ascending=False).head(top_n).reset_index(drop=True)


def fech_sir_por_causa_sir(df: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    """Volume não assertivo por causa SIR."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_CAUSA_SIR
    nao_asser = df[df['NAO_ASSERTIVO'] > 0] if 'NAO_ASSERTIVO' in df.columns else df[df['ASSERTIVO'] == 0]
    if nao_asser.empty or FECH_SIR_COL_CAUSA_SIR not in nao_asser.columns:
        return pd.DataFrame()
    vol_col = 'NAO_ASSERTIVO' if 'NAO_ASSERTIVO' in nao_asser.columns else 'ASSERTIVO'
    g = nao_asser.groupby(FECH_SIR_COL_CAUSA_SIR)[vol_col].sum().reset_index()
    g.columns = ['Causa SIR', 'Não Assertivo']
    g['Não Assertivo'] = g['Não Assertivo'].astype(int)
    return g.sort_values('Não Assertivo', ascending=False).head(top_n).reset_index(drop=True)


def fech_sir_por_regional(df: pd.DataFrame) -> pd.DataFrame:
    """Assertividade por regional."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_REGIONAL, FECH_SIR_COL_VOLUME
    if FECH_SIR_COL_REGIONAL not in df.columns:
        return pd.DataFrame()
    g = df.groupby(FECH_SIR_COL_REGIONAL).agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Assertivos=('ASSERTIVO', 'sum'),
    ).reset_index().rename(columns={FECH_SIR_COL_REGIONAL: 'Regional'})
    g['Assertividade_Pct'] = (g['Assertivos'] / g['Volume'] * 100).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


def fech_sir_por_demanda(df: pd.DataFrame) -> pd.DataFrame:
    """Assertividade por tipo de demanda."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_DEMANDA, FECH_SIR_COL_VOLUME
    if FECH_SIR_COL_DEMANDA not in df.columns:
        return pd.DataFrame()
    g = df.groupby(FECH_SIR_COL_DEMANDA).agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Assertivos=('ASSERTIVO', 'sum'),
    ).reset_index().rename(columns={FECH_SIR_COL_DEMANDA: 'Demanda'})
    g['Volume'] = g['Volume'].astype(int)
    g['Assertivos'] = g['Assertivos'].astype(int)
    g['Assertividade_Pct'] = (g['Assertivos'] / g['Volume'] * 100).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


def fech_sir_por_dia(df: pd.DataFrame) -> pd.DataFrame:
    """Evolução diária de assertividade."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_DIA, FECH_SIR_COL_VOLUME
    if FECH_SIR_COL_DIA not in df.columns:
        return pd.DataFrame()
    df2 = df.copy()
    df2['_dia'] = pd.to_numeric(df2[FECH_SIR_COL_DIA], errors='coerce')
    g = df2.groupby('_dia').agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Assertivos=('ASSERTIVO', 'sum'),
    ).reset_index().rename(columns={'_dia': 'Dia'})
    g['Assertividade_Pct'] = (g['Assertivos'] / g['Volume'] * 100).round(1)
    return g.sort_values('Dia').reset_index(drop=True)


def fech_sir_por_grupo(df: pd.DataFrame) -> pd.DataFrame:
    """Assertividade por IN_GRUPO (subgrupos dentro da Regional)."""
    if df.empty:
        return pd.DataFrame()
    from src.config import FECH_SIR_COL_GRUPO, FECH_SIR_COL_VOLUME
    if FECH_SIR_COL_GRUPO not in df.columns:
        return pd.DataFrame()
    g = df.groupby(FECH_SIR_COL_GRUPO).agg(
        Volume=(FECH_SIR_COL_VOLUME, 'sum'),
        Assertivos=('ASSERTIVO', 'sum'),
    ).reset_index().rename(columns={FECH_SIR_COL_GRUPO: 'Grupo'})
    g['Volume'] = g['Volume'].astype(int)
    g['Assertivos'] = g['Assertivos'].astype(int)
    g['Assertividade_Pct'] = (g['Assertivos'] / g['Volume'] * 100).round(1)
    return g.sort_values('Volume', ascending=False).reset_index(drop=True)


# =====================================================
# CHAT TOA — Loader e processadores
# =====================================================

def load_chat_toa(uploaded_file) -> pd.DataFrame:
    """
    Lê a planilha Analítico TOA Chat e retorna DataFrame processado.
    Filtra por:
      - Regional Leste (FECHAMENTO_FILA contains "RJO")
      - ANOMES mais recente
      - Equipe monitorada (ALL_TRACKED_IDS)

    Usa a coluna INDICADOR_TMA_DENTRO da planilha (0 ou 1) para determinar
    volume e aderência — sem recálculos de vida do chat ou TME.
    """
    from src.config import (
        CHAT_TOA_SHEET, CHAT_TOA_HEADER_ROW,
        CHAT_TOA_COL_INICIO, CHAT_TOA_COL_FIM, CHAT_TOA_COL_FIM_ANAL,
        CHAT_TOA_COL_ANOMES, CHAT_TOA_COL_LOGIN, CHAT_TOA_COL_FILA,
        CHAT_TOA_COL_IND_TMA,
        ALL_TRACKED_IDS, BASE_EQUIPE, COORD_ANALYSTS_NAMES, LOGIN_ALIASES,
    )

    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    df = pd.read_excel(uploaded_file, sheet_name=CHAT_TOA_SHEET, header=CHAT_TOA_HEADER_ROW)

    if CHAT_TOA_COL_LOGIN not in df.columns:
        return pd.DataFrame()

    # Normalizar login
    df[CHAT_TOA_COL_LOGIN] = (
        df[CHAT_TOA_COL_LOGIN].astype(str).str.strip().str.upper()
        .map(lambda x, a=LOGIN_ALIASES: a.get(x, x))
    )

    # ANOMES mais recente
    if CHAT_TOA_COL_ANOMES in df.columns:
        df[CHAT_TOA_COL_ANOMES] = pd.to_numeric(df[CHAT_TOA_COL_ANOMES], errors="coerce")
        anomes_max = df[CHAT_TOA_COL_ANOMES].max()
        df = df[df[CHAT_TOA_COL_ANOMES] == anomes_max].copy()

    # Filtro regional Leste
    if CHAT_TOA_COL_FILA in df.columns:
        df[CHAT_TOA_COL_FILA] = df[CHAT_TOA_COL_FILA].astype(str)
        df = df[df[CHAT_TOA_COL_FILA].str.contains("RJO", na=False)].copy()

    if df.empty:
        return pd.DataFrame()

    # Filtro equipe
    df = df[df[CHAT_TOA_COL_LOGIN].isin(ALL_TRACKED_IDS)].copy()
    if df.empty:
        return pd.DataFrame()

    # Converter datas (usadas apenas para eixos de evolução diária)
    for c in [CHAT_TOA_COL_INICIO, CHAT_TOA_COL_FIM, CHAT_TOA_COL_FIM_ANAL]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    # Volume e aderência vêm direto da planilha via INDICADOR_TMA_DENTRO.
    # Aceita apenas valores 0 e 1 (descarta fracionários e NaN).
    if CHAT_TOA_COL_IND_TMA in df.columns:
        df["VOL_TMA"] = df[CHAT_TOA_COL_IND_TMA].isin([0, 1]).astype(int)
        df["TMA_DENTRO"] = ((df["VOL_TMA"] == 1) & (df[CHAT_TOA_COL_IND_TMA] == 1)).astype(int)
    else:
        df["VOL_TMA"] = 0
        df["TMA_DENTRO"] = 0

    # Merge com nome e setor (via BASE_EQUIPE; fallback via COORD_ANALYSTS_NAMES)
    df = df.merge(
        BASE_EQUIPE[["Matricula", "Nome", "Setor"]],
        left_on=CHAT_TOA_COL_LOGIN, right_on="Matricula", how="left"
    ).drop(columns="Matricula", errors="ignore")
    df["Nome"] = df["Nome"].fillna(
        df[CHAT_TOA_COL_LOGIN].map(COORD_ANALYSTS_NAMES)
    ).fillna(df[CHAT_TOA_COL_LOGIN])
    df["Setor"] = df["Setor"].fillna("")

    # DATA_DIA para evolução diária
    df["DATA_DIA"] = df[CHAT_TOA_COL_INICIO].dt.normalize()

    return df


def chat_toa_kpis_gerais(df: pd.DataFrame) -> dict:
    """KPIs gerais: volume TMA, aderentes, aderência %, TMA médio (min)."""
    if df.empty:
        return {}
    from src.config import CHAT_TOA_COL_MINUTOS_TMA
    vol_tma = int(df["VOL_TMA"].sum())
    tma_ader = int(df["TMA_DENTRO"].sum())
    tma_medio = df.loc[df["VOL_TMA"] == 1, CHAT_TOA_COL_MINUTOS_TMA].mean() if vol_tma > 0 else None
    return {
        "vol_tma": vol_tma,
        "tma_aderentes": tma_ader,
        "tma_pct": round(tma_ader / vol_tma * 100, 1) if vol_tma > 0 else 0.0,
        "tma_medio_min": round(float(tma_medio), 1) if tma_medio is not None and not pd.isna(tma_medio) else None,
    }


def chat_toa_por_analista(df: pd.DataFrame) -> pd.DataFrame:
    """Ranking de aderência TMA por analista."""
    if df.empty:
        return pd.DataFrame()
    from src.config import CHAT_TOA_COL_LOGIN, CHAT_TOA_COL_MINUTOS_TMA
    g = df.groupby([CHAT_TOA_COL_LOGIN, "Nome", "Setor"]).agg(
        Vol_TMA=("VOL_TMA", "sum"),
        TMA_Aderentes=("TMA_DENTRO", "sum"),
    ).reset_index().rename(columns={CHAT_TOA_COL_LOGIN: "Login"})
    g["TMA_Pct"] = (g["TMA_Aderentes"] / g["Vol_TMA"] * 100).where(g["Vol_TMA"] > 0, 0).round(1)
    # TMA médio por analista
    tma_med = {}
    for login, grp in df.groupby(CHAT_TOA_COL_LOGIN):
        tma_sub = grp.loc[grp["VOL_TMA"] == 1, CHAT_TOA_COL_MINUTOS_TMA]
        tma_med[login] = tma_sub.mean() if not tma_sub.empty else None
    g["TMA_Medio_Min"] = g["Login"].map(tma_med).round(2)
    return g.sort_values("Vol_TMA", ascending=False).reset_index(drop=True)


def chat_toa_por_hora(df: pd.DataFrame) -> pd.DataFrame:
    """Distribuição de volume e aderência TMA por hora do dia."""
    if df.empty:
        return pd.DataFrame()
    from src.config import CHAT_TOA_COL_HORA
    sub = df[df["VOL_TMA"] == 1].copy()
    if sub.empty:
        return pd.DataFrame()
    g = sub.groupby(CHAT_TOA_COL_HORA).agg(
        Volume=("VOL_TMA", "sum"),
        TMA_Aderentes=("TMA_DENTRO", "sum"),
    ).reset_index().rename(columns={CHAT_TOA_COL_HORA: "Hora"})
    g["TMA_Pct"] = (g["TMA_Aderentes"] / g["Volume"] * 100).round(1)
    g["Hora"] = g["Hora"].astype(int)
    return g.sort_values("Hora").reset_index(drop=True)


def chat_toa_por_fila(df: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    """Aderência TMA por fila de fechamento (top N por volume)."""
    if df.empty:
        return pd.DataFrame()
    from src.config import CHAT_TOA_COL_FILA
    sub = df[df["VOL_TMA"] == 1]
    if sub.empty or CHAT_TOA_COL_FILA not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(CHAT_TOA_COL_FILA).agg(
        Volume=("VOL_TMA", "sum"),
        TMA_Aderentes=("TMA_DENTRO", "sum"),
    ).reset_index().rename(columns={CHAT_TOA_COL_FILA: "Fila"})
    g["TMA_Pct"] = (g["TMA_Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).head(top_n).reset_index(drop=True)


def chat_toa_evolucao_diaria(df: pd.DataFrame) -> pd.DataFrame:
    """Evolução diária de volume e aderência TMA."""
    if df.empty or "DATA_DIA" not in df.columns:
        return pd.DataFrame()
    sub = df.dropna(subset=["DATA_DIA"])
    g = sub.groupby("DATA_DIA").agg(
        Vol_TMA=("VOL_TMA", "sum"),
        TMA_Aderentes=("TMA_DENTRO", "sum"),
    ).reset_index().rename(columns={"DATA_DIA": "Data"})
    g["TMA_Pct"] = (g["TMA_Aderentes"] / g["Vol_TMA"] * 100).where(g["Vol_TMA"] > 0, 0).round(1)
    return g.sort_values("Data").reset_index(drop=True)


def chat_toa_por_tipo_fila(df: pd.DataFrame) -> pd.DataFrame:
    """Aderência TMA por tipo de fila (FO EMP, RF, FO RES…)."""
    if df.empty:
        return pd.DataFrame()
    from src.config import CHAT_TOA_COL_FILA_TIPO
    sub = df[df["VOL_TMA"] == 1]
    if sub.empty or CHAT_TOA_COL_FILA_TIPO not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(CHAT_TOA_COL_FILA_TIPO).agg(
        Volume=("VOL_TMA", "sum"),
        TMA_Aderentes=("TMA_DENTRO", "sum"),
    ).reset_index().rename(columns={CHAT_TOA_COL_FILA_TIPO: "Tipo Fila"})
    g["TMA_Pct"] = (g["TMA_Aderentes"] / g["Volume"] * 100).round(1)
    return g.sort_values("Volume", ascending=False).reset_index(drop=True)
